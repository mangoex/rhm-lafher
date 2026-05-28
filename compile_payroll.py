import openpyxl
import os
import datetime

def compile_payroll(input_path, output_path, config_incidences):
    """
    Lee el archivo de nómina base, aplica incidencias dinámicamente y escribe 
    fórmulas de Excel personalizadas por colaborador según sus condiciones.
    """
    if not os.path.exists(input_path):
        print(f"Error: El archivo de entrada no existe en {input_path}")
        return

    # Cargamos el libro manteniendo las fórmulas existentes
    print(f"Cargando plantilla base: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=False)
    ws = wb.active

    # La UMA está en R2 (columna R, renglón 2), pero las variables principales 
    # de los colaboradores inician en el renglón 6 hasta el 20.
    
    # 1. Buscamos la fila de cabeceras y determinamos el rango de colaboradores
    start_row = 6
    end_row = 20  # Según nuestro análisis, los 15 colaboradores están de la 6 a la 20
    
    print("\n--- INICIANDO COMPILACIÓN DE PRENÓMINA ---")
    
    for row in range(start_row, end_row + 1):
        # Leemos el código del colaborador (Columna B) y Nombre (Columna D)
        cod_cell = ws.cell(row=row, column=2)
        name_cell = ws.cell(row=row, column=4)
        
        cod = str(cod_cell.value).strip() if cod_cell.value is not None else None
        nombre = name_cell.value
        
        if not nombre:
            continue
            
        print(f"Procesando Fila {row} | Cód. {cod} | Colaborador: {nombre}")
        
        # 2. Obtenemos las incidencias parametrizadas para este colaborador
        incidencia = config_incidences.get(cod, {"faltas": 0, "descuento_adicional": 0, "observaciones": "OK"})
        faltas = incidencia.get("faltas", 0)
        descuento_adicional = incidencia.get("descuento_adicional", 0)
        observaciones = incidencia.get("observaciones", "OK")
        
        # 3. Escribir campos de control
        # Columna AD (Porcentaje de descuento adicional - opcional)
        ws.cell(row=row, column=30).value = 0 
        
        # Columna AE (Deducciones / Descuento Adicional, ej. abono a carro o préstamos)
        ws.cell(row=row, column=31).value = descuento_adicional
        
        # Columna AH (Observaciones de nómina)
        ws.cell(row=row, column=34).value = observaciones

        # 4. Inyección de Fórmulas Dinámicas Independientes por Fila
        # Columna AB (Sueldo Bruto Mensual) = SUM(U_row:AA_row)
        ws.cell(row=row, column=28).value = f"=SUM(U{row}:AA{row})"
        
        # Columna AC (Sueldo Bruto Quincenal base) = AB_row / 2
        ws.cell(row=row, column=29).value = f"=AB{row}/2"
        
        # Columna AF (Sueldo Bruto Mensual 2 - después de deducción mensual) = AB_row - AE_row
        ws.cell(row=row, column=32).value = f"=AB{row}-AE{row}"
        
        # Columna AG (Sueldo Bruto Quincenal 2 - Aplicando descuento proporcional por Faltas)
        # Si tiene faltas, dividimos el bruto mensual neto (AF) / 2 (quincena) / 15 (días del periodo)
        # y multiplicamos por los días laborados reales (15 - faltas).
        if faltas > 0:
            dias_laborados = 15 - faltas
            # Inyectamos una fórmula adaptada a esta fila por sus faltas
            ws.cell(row=row, column=33).value = f"=AF{row}/2/15*{dias_laborados}"
            print(f"  -> Inyectada fórmula de falta ({faltas} días): =AF{row}/2/15*{dias_laborados}")
        else:
            # Fórmula normal
            ws.cell(row=row, column=33).value = f"=AF{row}/2"
            print(f"  -> Inyectada fórmula normal: =AF{row}/2")
            
        # Columna AJ (Descuento quincenal acumulado) = AC_row - AG_row
        ws.cell(row=row, column=36).value = f"=AC{row}-AG{row}"

    # 5. Asegurar sumatorias totales en el renglón 21
    # Renglón 21 es la fila de sumatorias totales
    totals_row = 21
    # Columnas que se deben sumarizar verticalmente
    columns_to_sum = [
        (21, "U"), (22, "V"), (23, "W"), (24, "X"), (25, "Y"), (26, "Z"), (27, "AA"), 
        (28, "AB"), (31, "AE"), (32, "AF"), (36, "AJ")
    ]
    for col_idx, letter in columns_to_sum:
        ws.cell(row=totals_row, column=col_idx).value = f"=SUM({letter}6:{letter}20)"
        
    # AC21 (Bruto quincenal total) = AF21 / 2
    ws.cell(row=totals_row, column=29).value = f"=AF{totals_row}/2"
    # AG21 (Sueldo quincenal neto total) = SUM(AG6:AG20)
    ws.cell(row=totals_row, column=33).value = f"=SUM(AG6:AG20)"
    
    print("\nSumatorias de totales inyectadas en la Fila 21.")

    # Guardamos el archivo final con un nuevo nombre
    wb.save(output_path)
    print(f"\n¡Éxito! Archivo de nómina compilado guardado en: {output_path}")


if __name__ == "__main__":
    input_file = r"c:\Users\Miguel Gonzalez\Downloads\RHM\Nomina ciega.xlsx"
    output_file = r"c:\Users\Miguel Gonzalez\Downloads\RHM\Nomina_Compilada.xlsx"
    
    # Simulación de Incidencias enviadas por la Web UI:
    # Llave: Código del Colaborador en Excel (B6:B20)
    mock_incidences = {
        "190": { # Empleado 1
            "faltas": 1, 
            "descuento_adicional": 0.0,
            "observaciones": "DESCONTAR 1 DIA X RETARDOS (AUTOMATICO)"
        },
        "171": { # Empleado 2
            "faltas": 2, 
            "descuento_adicional": 0.0,
            "observaciones": "DESCONTAR 2 DIAS X INASISTENCIA JUSTIFICADA"
        },
        "108": { # Empleado 3 (Asimilado con observaciones de préstamo en Excel)
            "faltas": 0, 
            "descuento_adicional": 6244.18, # Descuento quincenal del préstamo de observaciones
            "observaciones": "DEDUCCION MENSUAL DE PRESTAMO 1 DE 5 ($6,244.18)"
        },
        "081": { # Empleado 4
            "faltas": 0, 
            "descuento_adicional": 0.0,
            "observaciones": "OK, SIN NOVEDAD"
        }
    }
    
    compile_payroll(input_file, output_file, mock_incidences)

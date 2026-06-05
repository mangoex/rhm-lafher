import os
import shutil
import datetime
import openpyxl
from server import (
    load_schema,
    check_and_heal_schema,
    recompile_active_period_incidences,
    parse_date_robust,
    parse_period_dates,
    clean_employee_id
)

def run_tests():
    print("=== STARTING VACATION TRACKING LOGIC TESTS ===")
    
    # 1. Setup copy of active database
    test_db = "test_vacation.xlsx"
    
    print(f"Creating mock database for test...")
    # Create a mock spreadsheet for testing
    wb = openpyxl.Workbook()
    ws = wb.active
    # Setup basic headers row 5
    ws.cell(row=5, column=1).value = "No."
    ws.cell(row=5, column=2).value = "Cod."
    ws.cell(row=5, column=4).value = "NOMBRE COMPLETO"
    ws.cell(row=5, column=9).value = "FECHA DE INGRESO"
    ws.cell(row=5, column=10).value = "AÑOS DE LABORES"
    ws.cell(row=5, column=14).value = "FACTOR INTEGRACION IMSS"
    
    # Row 6: Employee 1
    ws.cell(row=6, column=1).value = "1"
    ws.cell(row=6, column=2).value = "101"
    ws.cell(row=6, column=4).value = "Juan Perez"
    ws.cell(row=6, column=9).value = datetime.date(2024, 5, 10)
    
    # Row 7: Employee 2
    ws.cell(row=7, column=1).value = "2"
    ws.cell(row=7, column=2).value = "102"
    ws.cell(row=7, column=4).value = "Maria Lopez"
    ws.cell(row=7, column=9).value = datetime.date(2020, 8, 15)
    
    # Totals row
    ws.cell(row=8, column=4).value = "TOTAL"
    wb.save(test_db)
    wb.close()
        
    schema = load_schema()
    # Temporarily override db_path in schema for testing
    schema["db_path"] = test_db
    
    # 2. Test ensure_vacation_columns_in_excel
    print("\n[Test 1] Testing ensure_vacation_columns_in_excel...")
    from server import ensure_vacation_columns_in_excel
    ensure_vacation_columns_in_excel(test_db, schema)
    
    wb = openpyxl.load_workbook(test_db, data_only=False)
    ws = wb.active
    
    headers_row = 5 # typically 5
    h_tot = ws.cell(row=headers_row, column=38).value
    h_tom = ws.cell(row=headers_row, column=39).value
    h_res = ws.cell(row=headers_row, column=40).value
    
    print(f"Column 38 Header: {h_tot}")
    print(f"Column 39 Header: {h_tom}")
    print(f"Column 40 Header: {h_res}")
    
    assert "Vacaciones Totales" in str(h_tot), "Column 38 header mismatch!"
    assert "Vacaciones Tomadas" in str(h_tom), "Column 39 header mismatch!"
    assert "Vacaciones Restantes" in str(h_res), "Column 40 header mismatch!"
    print("[OK] Test 1 Passed: Vacation columns verified in Excel headers!")
    
    # 3. Create sample incidences for test_db
    # We will log:
    # Employee 101 (Ingreso: 2024-05-10):
    # - Anniversary 1 cycle (2025-05-10 to 2026-05-09). Earned 12 days.
    # - Active period: "16 al 30 de Abr 2026". The corte date is 2026-04-30.
    #   Antigüedad: ~1.97 years, completed_years = 1. Cycle: 2025-05-10 to 2026-05-09.
    # - Incidence on 2026-01-10: 3 days of vacation
    # - Incidence on 2026-04-20 (active period): 2 days of vacation
    # Employee 102 (Ingreso: 2020-08-15):
    # - Anniversary 5 cycle (2025-08-15 to 2026-08-14). Earned 20 days.
    # - Incidence on 2025-10-01: 5 days of vacation
    # - Incidence on 2026-04-18 (active period): 1 day of vacation
    
    print("\n[Test 2] Setting up mock Incidencias sheet...")
    if "Incidencias" in wb.sheetnames:
        wb.remove(wb["Incidencias"])
    ws_inc = wb.create_sheet("Incidencias")
    headers = [
        "Fecha", "Código", "Nombre", "Faltas", "Retardos", "Vacaciones", 
        "Descuento Adicional", "Puntualidad", "Asistencia", "Observaciones",
        "Forzar Asistencia", "Forzar Puntualidad", "Forzar Vales", "Ajuste Vales", "Ajuste Fondo Ahorro"
    ]
    ws_inc.append(headers)
    
    # Incidencias for Employee 101
    ws_inc.append(["2026-01-10", "101", "Juan Perez", 0, 0, 3, 0.0, "SI", "SI", "Previas", "NO", "NO", "NO", None, None])
    ws_inc.append(["2026-04-20", "101", "Juan Perez", 0, 0, 2, 0.0, "SI", "SI", "En activo", "NO", "NO", "NO", None, None])
    
    # Incidencias for Employee 102
    ws_inc.append(["2025-10-01", "102", "Maria Lopez", 0, 0, 5, 0.0, "SI", "SI", "Previas", "NO", "NO", "NO", None, None])
    ws_inc.append(["2026-04-18", "102", "Maria Lopez", 0, 0, 1, 0.0, "SI", "SI", "En activo", "NO", "NO", "NO", None, None])
    
    # An out of range incidence for Employee 101 (belonging to previous cycle 2024-05-10 to 2025-05-09)
    # This should NOT be counted in the current cycle tomadas!
    ws_inc.append(["2025-02-15", "101", "Juan Perez", 0, 0, 4, 0.0, "SI", "SI", "Ciclo anterior", "NO", "NO", "NO", None, None])
    
    wb.save(test_db)
    wb.close()
    print("[OK] Test 2 Passed: Mock Incidencias sheet saved!")

    # 4. Test recompile_active_period_incidences
    print("\n[Test 3] Running recompile_active_period_incidences...")
    wb_test = openpyxl.load_workbook(test_db, data_only=False)
    # We will use active period "16 al 30 de Abr 2026"
    schema["period"] = "16 al 30 Abr 2026"
    
    recompile_active_period_incidences(wb_test, schema)
    wb_test.save(test_db)
    wb_test.close()
    
    # Verify calculated values in Hoja1
    print("\n[Test 4] Verifying recalculated values...")
    wb_read = openpyxl.load_workbook(test_db, data_only=True)
    ws_read = wb_read.active
    
    # Employee 101 is row 6
    name_101 = ws_read.cell(row=6, column=4).value
    id_101 = ws_read.cell(row=6, column=2).value
    ingreso_101 = ws_read.cell(row=6, column=9).value
    antiguedad_101 = ws_read.cell(row=6, column=10).value
    vac_tot_101 = ws_read.cell(row=6, column=38).value
    vac_tom_101 = ws_read.cell(row=6, column=39).value
    
    print(f"\nEmployee {id_101} ({name_101}):")
    print(f"  Fecha de Ingreso: {ingreso_101}")
    print(f"  Antigüedad calculada (Años): {antiguedad_101}")
    print(f"  Vacaciones Totales (Derecho): {vac_tot_101}")
    print(f"  Vacaciones Tomadas en ciclo: {vac_tom_101}")
    
    # Assert Employee 101:
    # Ingreso: 2024-05-10. Corte: 2026-04-30. Years of service: 1.97 years
    # Completed years: 1. LFT derecho: 12 days.
    # Vacaciones tomadas in cycle [2025-05-10, 2026-05-09):
    # - 2026-01-10: 3 days (in range)
    # - 2026-04-20: 2 days (in range)
    # - 2025-02-15: 4 days (NOT in range, previous cycle)
    # Total Tomadas should be 3 + 2 = 5 days.
    
    assert abs(antiguedad_101 - 1.97) < 0.02, f"Antigüedad mismatch! Expected ~1.97, got {antiguedad_101}"
    assert int(vac_tot_101) == 12, f"Vacaciones Totales mismatch! Expected 12, got {vac_tot_101}"
    assert int(vac_tom_101) == 5, f"Vacaciones Tomadas mismatch! Expected 5, got {vac_tom_101}"
    print("[OK] Employee 101 calculations verified!")
    
    # Employee 102 is row 7
    name_102 = ws_read.cell(row=7, column=4).value
    id_102 = ws_read.cell(row=7, column=2).value
    ingreso_102 = ws_read.cell(row=7, column=9).value
    antiguedad_102 = ws_read.cell(row=7, column=10).value
    vac_tot_102 = ws_read.cell(row=7, column=38).value
    vac_tom_102 = ws_read.cell(row=7, column=39).value
    
    print(f"\nEmployee {id_102} ({name_102}):")
    print(f"  Fecha de Ingreso: {ingreso_102}")
    print(f"  Antigüedad calculada (Años): {antiguedad_102}")
    print(f"  Vacaciones Totales (Derecho): {vac_tot_102}")
    print(f"  Vacaciones Tomadas en ciclo: {vac_tom_102}")
    
    # Assert Employee 102:
    # Ingreso: 2020-08-15. Corte: 2026-04-30. Years of service: 5.71 years.
    # Completed years: 5. LFT derecho: 20 days.
    # Vacaciones tomadas in cycle [2025-08-15, 2026-08-14):
    # - 2025-10-01: 5 days (in range)
    # - 2026-04-18: 1 day (in range)
    # Total Tomadas should be 5 + 1 = 6 days.
    
    assert abs(antiguedad_102 - 5.71) < 0.02, f"Antigüedad mismatch! Expected ~5.71, got {antiguedad_102}"
    assert int(vac_tot_102) == 20, f"Vacaciones Totales mismatch! Expected 20, got {vac_tot_102}"
    assert int(vac_tom_102) == 6, f"Vacaciones Tomadas mismatch! Expected 6, got {vac_tom_102}"
    print("[OK] Employee 102 calculations verified!")
    
    # Verify remaining vacations and Factor de Integracion formula
    wb_f = openpyxl.load_workbook(test_db, data_only=False)
    ws_f = wb_f.active
    
    fi_formula_101 = ws_f.cell(row=6, column=14).value
    fi_formula_102 = ws_f.cell(row=7, column=14).value
    formula_101 = ws_f.cell(row=6, column=40).value
    formula_102 = ws_f.cell(row=7, column=40).value
    
    print(f"\nFactor de Integracion Formula Row 6: {fi_formula_101}")
    print(f"Factor de Integracion Formula Row 7: {fi_formula_102}")
    print(f"Remaining Vacations Formula Row 6: {formula_101}")
    print(f"Remaining Vacations Formula Row 7: {formula_102}")
    
    assert fi_formula_101.startswith("=1 + "), f"Factor de Integración formula row 6 mismatch: {fi_formula_101}"
    assert fi_formula_102.startswith("=1 + "), f"Factor de Integración formula row 7 mismatch: {fi_formula_102}"
    assert formula_101 == "=AL6-AM6", "Formula mismatch in row 6!"
    assert formula_102 == "=AL7-AM7", "Formula mismatch in row 7!"
    print("[OK] Remaining vacations and Factor de Integración formulas verified!")
    
    wb_read.close()
    wb_f.close()
    
    # Cleanup test file
    if os.path.exists(test_db):
        os.remove(test_db)
        
    print("\n=== ALL TESTS PASSED SUCCESSFULLY! ===")

if __name__ == "__main__":
    run_tests()

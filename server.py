import http.server
import socketserver
import json
import os
import openpyxl
import urllib.parse
import urllib.request
from datetime import datetime
import traceback
import os

PORT = 8000
import sys

# Resolve paths for PyInstaller standalone executables
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    STATIC_DIR = getattr(sys, '_MEIPASS', BASE_DIR)
    
    # Save configuration and default database in user folder to ensure writability on macOS/Windows
    CONFIG_DIR = os.path.expanduser("~/.rhm_prenomina")
    try:
        os.makedirs(CONFIG_DIR, exist_ok=True)
    except Exception as e:
        print(f"Error creating user config directory: {e}")
        CONFIG_DIR = BASE_DIR
    SCHEMA_PATH = os.path.join(CONFIG_DIR, "schema.json")
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    STATIC_DIR = BASE_DIR
    
    # Test if CONFIG_DIR (BASE_DIR) is writable; if not, fallback to a writable folder
    CONFIG_DIR = BASE_DIR
    test_file = os.path.join(CONFIG_DIR, ".write_test")
    try:
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
    except Exception:
        CONFIG_DIR = os.path.expanduser("~/.rhm_prenomina")
        try:
            os.makedirs(CONFIG_DIR, exist_ok=True)
        except Exception:
            CONFIG_DIR = "/tmp"
            
    SCHEMA_PATH = os.path.join(CONFIG_DIR, "schema.json")

# Support cloud persistent storage volume via environment variable
env_config_dir = os.environ.get("RHM_DATA_DIR")
if env_config_dir:
    proposed_dir = os.path.abspath(env_config_dir)
    is_writable = False
    try:
        os.makedirs(proposed_dir, exist_ok=True)
        # Test write capability in proposed directory
        test_file = os.path.join(proposed_dir, ".rhm_write_test")
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
        is_writable = True
    except Exception as e:
        print(f"Error checking writability of RHM_DATA_DIR ({proposed_dir}): {e}")
        
    if is_writable:
        CONFIG_DIR = proposed_dir
        SCHEMA_PATH = os.path.join(CONFIG_DIR, "schema.json")
    else:
        print(f"RHM_DATA_DIR ({proposed_dir}) is not writable. Falling back to default CONFIG_DIR: {CONFIG_DIR}")

import shutil
import hashlib
import secrets
import time
import threading

EXCEL_LOCK = threading.Lock()

USERS_FILE = os.path.join(CONFIG_DIR, "users.json")
COMPANIES_FILE = os.path.join(CONFIG_DIR, "companies.json")
SECRETS_PATH = os.path.join(CONFIG_DIR, "secrets.json")
SESSIONS = {}  # token -> {"username": username, "role": role, "expiry": timestamp}

def load_secrets():
    if not os.path.exists(SECRETS_PATH):
        return {}
    try:
        with open(SECRETS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_secrets(data):
    try:
        with open(SECRETS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving secrets: {e}")

def hash_password(password, salt):
    return hashlib.sha256((password + salt).encode('utf-8')).hexdigest()

def load_users():
    if not os.path.exists(USERS_FILE):
        salt = secrets.token_hex(8)
        hashed = hash_password("adminpassword123", salt)
        users = {
            "admin": {
                "username": "admin",
                "role": "admin",
                "salt": salt,
                "password": hashed
            }
        }
        save_users(users)
        return users
    try:
        with open(USERS_FILE, "r") as f:
            return json.load(f)
    except Exception as e:
        print("Error loading users:", e)
        return {}

def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=4)

def load_companies():
    if not os.path.exists(COMPANIES_FILE):
        default_seed_path = os.path.join(STATIC_DIR, "companies.json")
        if os.path.exists(default_seed_path):
            try:
                shutil.copyfile(default_seed_path, COMPANIES_FILE)
                with open(COMPANIES_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                pass
        
        # Fallback default seed
        companies = [
            { "id": "1", "nombre": "AHA", "razon_social": "AHA S.A. de C.V.", "regimen": "Régimen General de Ley Personas Morales", "prima_riesgo": 0.5432 },
            { "id": "2", "nombre": "BYRMAX", "razon_social": "BYRMAX S. de R.L. de C.V.", "regimen": "Régimen General de Ley Personas Morales", "prima_riesgo": 1.1345 },
            { "id": "3", "nombre": "CPI", "razon_social": "CPI Servicios Financieros S.A. de C.V.", "regimen": "Régimen General de Ley Personas Morales", "prima_riesgo": 0.5000 },
            { "id": "4", "nombre": "CPI/ ASIMIL", "razon_social": "CPI Asimilados y Servicios Integrales", "regimen": "Sueldos y Salarios / Asimilados", "prima_riesgo": 0.0000 },
            { "id": "5", "nombre": "LASO", "razon_social": "LASO Corporativo Jurídico", "regimen": "Régimen General de Ley Personas Morales", "prima_riesgo": 0.7500 },
            { "id": "6", "nombre": "FACTURA", "razon_social": "Facturación y Comisiones", "regimen": "Sueldos y Salarios / Asimilados", "prima_riesgo": 0.0000 }
        ]
        save_companies_data(companies)
        return companies
        
    try:
        with open(COMPANIES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print("Error loading companies:", e)
        return []

def save_companies_data(companies):
    try:
        with open(COMPANIES_FILE, "w", encoding="utf-8") as f:
            json.dump(companies, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print("Error saving companies:", e)

# Extract schema.json first if missing and frozen
if getattr(sys, 'frozen', False):
    bundled_schema = os.path.join(STATIC_DIR, "schema.json")
    if not os.path.exists(SCHEMA_PATH) and os.path.exists(bundled_schema):
        print(f"schema.json missing in config directory. Copying template to: {SCHEMA_PATH}")
        try:
            shutil.copyfile(bundled_schema, SCHEMA_PATH)
        except Exception as e:
            print("Error extracting template schema.json:", e)

def extract_default_payroll_rules(docx_path):
    if not os.path.exists(docx_path):
        return ""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        import io
        with open(docx_path, 'rb') as f:
            data_bytes = f.read()
        with zipfile.ZipFile(io.BytesIO(data_bytes)) as z:
            doc_xml = z.read('word/document.xml')
            root = ET.fromstring(doc_xml)
            paragraphs = []
            for elem in root.iter():
                tag_name = elem.tag.split('}')[-1]
                if tag_name == 'p':
                    p_text = []
                    for child in elem.iter():
                        child_tag = child.tag.split('}')[-1]
                        if child_tag == 't' and child.text:
                            p_text.append(child.text)
                    text_str = "".join(p_text).strip()
                    if text_str:
                        paragraphs.append(text_str)
            return "\n\n".join(paragraphs)
    except Exception as e:
        print("Error extracting default payroll rules from docx:", e)
        return ""

MONTHS_ES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12
}

def clean_employee_id(id_val):
    if id_val is None:
        return ""
    s = str(id_val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def parse_date_robust(val):
    if not val:
        return None
    import datetime
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(s[:10], fmt).date()
        except:
            pass
    # Regex fallback if parsed differently
    import re
    m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        try: return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except: pass
    m2 = re.match(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m2:
        try: return datetime.date(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
        except: pass
    return None

def calculate_vacation_days(years):
    if years < 1: return 0
    if years == 1: return 12
    if years == 2: return 14
    if years == 3: return 16
    if years == 4: return 18
    if years == 5: return 20
    if years <= 10: return 22
    if years <= 15: return 24
    if years <= 20: return 26
    if years <= 25: return 28
    if years <= 30: return 30
    return 20 + 2 * ((years - 1) // 5)

def parse_period_dates(period_str):
    import re
    import datetime
    m = re.match(r"(\d+)\s+al\s+(\d+)\s+(\w+)\s+(\d{4})", period_str.strip(), re.IGNORECASE)
    if m:
        d_start = int(m.group(1))
        d_end = int(m.group(2))
        month_str = m.group(3).lower()[:3]
        year = int(m.group(4))
        month = MONTHS_ES.get(month_str, 4)
        
        start_date = datetime.date(year, month, d_start)
        end_date = datetime.date(year, month, d_end)
        return start_date, end_date
    else:
        today = datetime.date.today()
        if today.day <= 15:
            return datetime.date(today.year, today.month, 1), datetime.date(today.year, today.month, 15)
        else:
            import calendar
            _, last_day = calendar.monthrange(today.year, today.month)
            return datetime.date(today.year, today.month, 16), datetime.date(today.year, today.month, last_day)

def parse_multipart(body_bytes, boundary):
    parts = body_bytes.split(b'--' + boundary.encode('utf-8'))
    result = {}
    for part in parts:
        if not part or part == b'\r\n' or part.startswith(b'--'):
            continue
        if b'\r\n\r\n' not in part:
            continue
        header_part, content = part.split(b'\r\n\r\n', 1)
        if content.endswith(b'\r\n'):
            content = content[:-2]
            
        headers = {}
        for line in header_part.split(b'\r\n'):
            if b':' in line:
                key, val = line.split(b':', 1)
                headers[key.decode('utf-8').strip().lower()] = val.decode('utf-8').strip()
                
        disp = headers.get('content-disposition', '')
        if 'form-data' in disp:
            name = None
            filename = None
            for item in disp.split(';'):
                item = item.strip()
                if item.startswith('name='):
                    name = item.split('=', 1)[1].strip('"')
                elif item.startswith('filename='):
                    filename = item.split('=', 1)[1].strip('"')
            
            if name:
                result[name] = {
                    'filename': filename,
                    'content': content,
                    'content-type': headers.get('content-type', '')
                }
    return result

def ensure_absolute_cell(coord):
    if not coord: return coord
    if "$" in coord: return coord
    import re
    m = re.match(r"([A-Z]+)([0-9]+)", coord)
    if m:
        return f"${m.group(1)}${m.group(2)}"
    return coord

def heal_incidences_sheet_if_needed(wb, schema):
    if "Incidencias" not in wb.sheetnames:
        return False
    ws_inc = wb["Incidencias"]
    if ws_inc.max_row < 1:
        return False
    
    # Read headers
    headers = [ws_inc.cell(row=1, column=c).value for c in range(1, ws_inc.max_column + 1)]
    headers = [h for h in headers if h is not None]
    
    # Check if Forzar Asistencia is already present
    if "Forzar Asistencia" in headers or "FORZAR ASISTENCIA" in [str(h).upper() for h in headers]:
        return False
        
    print("Migrating old Incidencias sheet to include override columns...")
    
    # Read all rows
    old_rows = []
    for r in range(2, ws_inc.max_row + 1):
        row_vals = [ws_inc.cell(row=r, column=c).value for c in range(1, ws_inc.max_column + 1)]
        old_rows.append(row_vals)
        
    # Recreate sheet
    wb.remove(ws_inc)
    ws_inc = wb.create_sheet("Incidencias")
    
    new_headers = [
        "Fecha", "CÃ³digo", "Nombre", "Faltas", "Retardos", "Vacaciones", 
        "Descuento Adicional", "Puntualidad", "Asistencia", "Observaciones",
        "Forzar Asistencia", "Forzar Puntualidad", "Forzar Vales", "Ajuste Vales", "Ajuste Fondo Ahorro"
    ]
    for col in schema.get("columns", []):
        if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
            new_headers.append(col.get("label") or col.get("header") or col.get("field"))
            
    ws_inc.append(new_headers)
    
    # Write back rows
    for row_vals in old_rows:
        new_row = []
        old_header_len = len(headers)
        if len(row_vals) < old_header_len:
            row_vals.extend([None] * (old_header_len - len(row_vals)))
            
        new_row.extend(row_vals[:10])
        # Forzar Asistencia (NO), Forzar Puntualidad (NO), Forzar Vales (NO), Ajuste Vales (None), Ajuste Fondo Ahorro (None)
        new_row.extend(["NO", "NO", "NO", None, None])
        new_row.extend(row_vals[10:])
        ws_inc.append(new_row)
        
    return True

def save_incidence_to_excel(wb, data):
    import datetime
    schema = check_and_heal_schema()
    
    # Check and migrate sheet if needed
    heal_incidences_sheet_if_needed(wb, schema)
    
    if "Incidencias" not in wb.sheetnames:
        ws_inc = wb.create_sheet("Incidencias")
        headers = [
            "Fecha", "CÃ³digo", "Nombre", "Faltas", "Retardos", "Vacaciones", 
            "Descuento Adicional", "Puntualidad", "Asistencia", "Observaciones",
            "Forzar Asistencia", "Forzar Puntualidad", "Forzar Vales", "Ajuste Vales", "Ajuste Fondo Ahorro"
        ]
        for col in schema.get("columns", []):
            if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                headers.append(col.get("label") or col.get("header") or col.get("field"))
        ws_inc.append(headers)
    else:
        ws_inc = wb["Incidencias"]
    
    date_str = data.get("date")
    if not date_str:
        # Use Mexico City time (UTC-6)
        tz_mex = datetime.timezone(datetime.timedelta(hours=-6))
        today_mex = datetime.datetime.now(tz_mex).date()
        date_str = today_mex.strftime("%Y-%m-%d")
    
    cod = clean_employee_id(data.get("id"))
    
    found_row = None
    for r in range(2, ws_inc.max_row + 1):
        r_date = ws_inc.cell(row=r, column=1).value
        r_cod = ws_inc.cell(row=r, column=2).value
        r_date_parsed = parse_date_robust(r_date)
        if r_date_parsed and r_cod:
            r_date_str = r_date_parsed.strftime("%Y-%m-%d")
            if r_date_str == date_str and clean_employee_id(r_cod) == cod:
                found_row = r
                break
            
    def parse_float_opt(val):
        if val is None or str(val).strip() == "":
            return None
        try:
            return float(val)
        except ValueError:
            return None

    row_values = [
        date_str,
        cod,
        data.get("nombre", ""),
        int(data.get("faltas", 0)),
        int(data.get("retardos", 0)),
        int(data.get("vacaciones", 0)),
        float(data.get("descuento_adicional", 0.0)),
        data.get("puntualidad", "SI"),
        data.get("asistencia", "SI"),
        data.get("observaciones", ""),
        data.get("forzar_asistencia", "NO"),
        data.get("forzar_puntualidad", "NO"),
        data.get("forzar_vales", "NO"),
        parse_float_opt(data.get("ajuste_vales")),
        parse_float_opt(data.get("ajuste_fondo_ahorro"))
    ]
    
    # Append dynamic fields from data
    for col in schema.get("columns", []):
        if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
            row_values.append(float(data.get(col.get("field"), 0.0)))
    
    if found_row:
        for col_idx, val in enumerate(row_values, 1):
            ws_inc.cell(row=found_row, column=col_idx).value = val
    else:
        ws_inc.append(row_values)


def recompile_active_period_incidences(wb, schema):
    ws = wb.active # Hoja1
    
    def add_years(d, years):
        try:
            return d.replace(year=d.year + years)
        except ValueError:
            return d.replace(year=d.year + years, day=28)
            
    def get_vacation_allowance(y):
        return calculate_vacation_days(y)
            
    # First, make sure Incidencias is migrated if needed
    heal_incidences_sheet_if_needed(wb, schema)
    
    period_str = schema.get("period", "16 al 30 Abr 2026")
    start_date, end_date = parse_period_dates(period_str)
    
    agg = {}
    all_incidences = []
    if "Incidencias" in wb.sheetnames:
        ws_inc = wb["Incidencias"]
        # Batch load all rows in memory (tuple format: 0-indexed)
        rows = list(ws_inc.iter_rows(min_row=2, max_row=ws_inc.max_row, values_only=True))
        
        for r_idx, row_cells in enumerate(rows, start=2):
            if not row_cells:
                continue
            
            # Ensure we have at least column 1 value (date_val)
            date_val = row_cells[0] if len(row_cells) >= 1 else None
            if not date_val:
                continue
                
            try:
                row_date = parse_date_robust(date_val)
                cod_val = row_cells[1] if len(row_cells) >= 2 else None
                c_id = clean_employee_id(cod_val)
                
                # Load for historical vacations check
                if row_date and c_id and len(row_cells) >= 6:
                    try:
                        vac_val = row_cells[5]
                        vac_days = int(vac_val or 0)
                    except:
                        vac_days = 0
                    if vac_days > 0:
                        all_incidences.append({
                            "id": c_id,
                            "date": row_date,
                            "vacaciones": vac_days
                        })
                
                if row_date and start_date <= row_date <= end_date:
                    if c_id not in agg:
                        agg[c_id] = {
                            "faltas": 0,
                            "retardos": 0,
                            "vacaciones": 0,
                            "descuento_adicional": 0.0,
                            "puntualidad": "SI",
                            "asistencia": "SI",
                            "forzar_asistencia": "NO",
                            "forzar_puntualidad": "NO",
                            "forzar_vales": "NO",
                            "ajuste_vales": None,
                            "ajuste_fondo_ahorro": None,
                            "observaciones": []
                        }
                        for col in schema.get("columns", []):
                            if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                                agg[c_id][col.get("field")] = 0.0
                                
                    # Column 4 (index 3)
                    faltas_val = row_cells[3] if len(row_cells) >= 4 else 0
                    agg[c_id]["faltas"] += int(faltas_val or 0)
                    
                    # Column 5 (index 4)
                    retardos_val = row_cells[4] if len(row_cells) >= 5 else 0
                    agg[c_id]["retardos"] += int(retardos_val or 0)
                    
                    # Column 6 (index 5)
                    vacaciones_val = row_cells[5] if len(row_cells) >= 6 else 0
                    agg[c_id]["vacaciones"] += int(vacaciones_val or 0)
                    
                    # Column 7 (index 6)
                    desc_val = row_cells[6] if len(row_cells) >= 7 else 0.0
                    agg[c_id]["descuento_adicional"] += float(desc_val or 0.0)
                    
                    # Column 8 (index 7)
                    punt_val = row_cells[7] if len(row_cells) >= 8 else None
                    if punt_val == "NO":
                        agg[c_id]["puntualidad"] = "NO"
                        
                    # Column 9 (index 8)
                    asist_val = row_cells[8] if len(row_cells) >= 9 else None
                    if asist_val == "NO":
                        agg[c_id]["asistencia"] = "NO"
                        
                    # Column 10 (index 9)
                    obs_val = row_cells[9] if len(row_cells) >= 10 else ""
                    obs = str(obs_val or "").strip()
                    if obs:
                        agg[c_id]["observaciones"].append(obs)
                        
                    # Overrides (columns 11 to 15)
                    max_cols = len(row_cells)
                    if max_cols >= 11 and row_cells[10] == "SI":
                        agg[c_id]["forzar_asistencia"] = "SI"
                    if max_cols >= 12 and row_cells[11] == "SI":
                        agg[c_id]["forzar_puntualidad"] = "SI"
                    if max_cols >= 13 and row_cells[12] == "SI":
                        agg[c_id]["forzar_vales"] = "SI"
                    
                    if max_cols >= 14:
                        aj_vales = row_cells[13]
                        if aj_vales is not None and str(aj_vales).strip() != "":
                            try:
                                agg[c_id]["ajuste_vales"] = float(aj_vales)
                            except ValueError:
                                pass
                                
                    if max_cols >= 15:
                        aj_fa = row_cells[14]
                        if aj_fa is not None and str(aj_fa).strip() != "":
                            try:
                                agg[c_id]["ajuste_fondo_ahorro"] = float(aj_fa)
                            except ValueError:
                                pass
                        
                    # Sum up dynamic deduction columns starting from index 16 (index 15 in 0-indexed)
                    col_idx = 16
                    for col in schema.get("columns", []):
                        if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                            if col_idx <= max_cols:
                                agg[c_id][col.get("field")] += float(row_cells[col_idx - 1] or 0.0)
                                col_idx += 1
            except Exception as e:
                print(f"Error compiling incidence row {r_idx}: {e}")
                
    nombre_col = get_field_index(schema, "nombre")
    id_col = get_field_index(schema, "id")
    descuento_col = get_field_index(schema, "descuento_adicional")
    observaciones_col = get_field_index(schema, "observaciones")
    neto_quincenal_col = get_field_index(schema, "neto_quincenal")
    bruto_mensual_neto_letter = get_field_letter(schema, "bruto_mensual_neto")
    sdi_letter = get_field_letter(schema, "sdi")
    puntualidad_col = get_field_index(schema, "puntualidad")
    asistencia_col = get_field_index(schema, "asistencia")
    vales_col = get_field_index(schema, "vales_despensa")
    fa_col = get_field_index(schema, "fondo_ahorro")
    
    ingreso_col = get_field_index(schema, "ingreso")
    antiguedad_col = get_field_index(schema, "antiguedad")
    fi_col = get_field_index(schema, "factor_integracion")
    vac_totales_col = get_field_index(schema, "vacaciones_totales")
    vac_tomadas_col = get_field_index(schema, "vacaciones_tomadas")
    vac_restantes_col = get_field_index(schema, "vacaciones_restantes")
    
    # Cells config
    uma_cell = ensure_absolute_cell(schema.get("uma_cell", "S3"))
    vales_pct_cell = ensure_absolute_cell(schema.get("vales_pct_cell", "P3"))
    dias_mes_cell = ensure_absolute_cell(schema.get("dias_mes_cell", "N3"))
    fa_pct_cell = ensure_absolute_cell(schema.get("fa_pct_cell", "L3"))
    
    headers_row = find_headers_row(ws)
    row = headers_row + 1
    
    while True:
        nombre_val = ws.cell(row=row, column=nombre_col).value
        cod_val = ws.cell(row=row, column=id_col).value
        
        if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
            break
        if nombre_val is None and cod_val is None:
            has_more = False
            for i in range(1, 4):
                n = ws.cell(row=row+i, column=nombre_col).value
                c = ws.cell(row=row+i, column=id_col).value
                if n or c:
                    has_more = True
            if not has_more:
                break
        
        if nombre_val:
            cod_id = clean_employee_id(cod_val) if cod_val is not None else f"TEMP_{row}"
            
            # Calculate and set dynamic calculations for vacations, antiquity and Factor de Integración
            if ingreso_col:
                ingreso_val = ws.cell(row=row, column=ingreso_col).value
                ingreso_dt = parse_date_robust(ingreso_val)
                baja_col = get_field_index(schema, "baja")
                baja_val = ws.cell(row=row, column=baja_col).value if baja_col else None
                baja_dt = parse_date_robust(baja_val)
                
                if ingreso_dt:
                    corte_date = baja_dt if baja_dt else end_date
                    diff_days = (corte_date - ingreso_dt).days
                    antiguedad_val = round(diff_days / 365.25, 4)
                    
                    if antiguedad_col:
                        ws.cell(row=row, column=antiguedad_col).value = antiguedad_val
                        
                    completed_years = int(diff_days / 365.25)
                    vac_derecho = get_vacation_allowance(completed_years)
                    
                    if vac_totales_col:
                        ws.cell(row=row, column=vac_totales_col).value = float(vac_derecho)
                        
                    last_anniversary = add_years(ingreso_dt, completed_years)
                    next_anniversary = add_years(ingreso_dt, completed_years + 1)
                    
                    vac_tomadas = 0
                    for inc in all_incidences:
                        if inc["id"] == cod_id and last_anniversary <= inc["date"] < next_anniversary:
                            vac_tomadas += inc["vacaciones"]
                            
                    if vac_tomadas_col:
                        ws.cell(row=row, column=vac_tomadas_col).value = float(vac_tomadas)
                        
                    if vac_restantes_col:
                        tot_letter = get_field_letter(schema, "vacaciones_totales")
                        tom_letter = get_field_letter(schema, "vacaciones_tomadas")
                        ws.cell(row=row, column=vac_restantes_col).value = f"={tot_letter}{row}-{tom_letter}{row}"
                        
                    if fi_col:
                        if not baja_dt:
                            fi = 1 + (15/365) + ((vac_derecho * 0.25) / 365)
                            ws.cell(row=row, column=fi_col).value = round(fi, 4)
                        else:
                            ws.cell(row=row, column=fi_col).value = 0.0
                else:
                    if antiguedad_col: ws.cell(row=row, column=antiguedad_col).value = 0.0
                    if vac_totales_col: ws.cell(row=row, column=vac_totales_col).value = 0.0
                    if vac_tomadas_col: ws.cell(row=row, column=vac_tomadas_col).value = 0.0
                    if vac_restantes_col: ws.cell(row=row, column=vac_restantes_col).value = 0.0
                    if fi_col: ws.cell(row=row, column=fi_col).value = 0.0

            # Reset defaults in Hoja1
            if descuento_col:
                ws.cell(row=row, column=descuento_col).value = None
            if observaciones_col:
                ws.cell(row=row, column=observaciones_col).value = None
            if neto_quincenal_col and bruto_mensual_neto_letter:
                ws.cell(row=row, column=neto_quincenal_col).value = f"={bruto_mensual_neto_letter}{row}/2"
            
            if puntualidad_col and sdi_letter:
                ws.cell(row=row, column=puntualidad_col).value = f"={sdi_letter}{row}*0.1*{dias_mes_cell}"
            if asistencia_col and sdi_letter:
                ws.cell(row=row, column=asistencia_col).value = f"={sdi_letter}{row}*0.1*{dias_mes_cell}"
            
            if vales_col:
                ws.cell(row=row, column=vales_col).value = f"={uma_cell}*({vales_pct_cell}/100)*{dias_mes_cell}"
            if fa_col:
                ws.cell(row=row, column=fa_col).value = f'=IF({get_field_letter(schema, "fondo_ahorro_activo")}{row}="SI", MIN({get_field_letter(schema, "sueldo_nominal")}{row}*({fa_pct_cell}/100), 1.3*{uma_cell}*{dias_mes_cell}), 0)'
                
            # Reset dynamic deductions in Hoja1
            for col in schema.get("columns", []):
                if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                    ws.cell(row=row, column=col.get("index")).value = None
                
            if cod_id in agg:
                emp_agg = agg[cod_id]
                faltas = emp_agg["faltas"]
                retardos = emp_agg["retardos"]
                desc = emp_agg["descuento_adicional"]
                obs_list = emp_agg["observaciones"]
                punt = emp_agg["puntualidad"]
                asist = emp_agg["asistencia"]
                
                forzar_asist = emp_agg.get("forzar_asistencia", "NO") == "SI"
                forzar_punt = emp_agg.get("forzar_puntualidad", "NO") == "SI"
                forzar_val = emp_agg.get("forzar_vales", "NO") == "SI"
                ajuste_vales = emp_agg.get("ajuste_vales")
                ajuste_fa = emp_agg.get("ajuste_fondo_ahorro")
                
                if neto_quincenal_col and bruto_mensual_neto_letter:
                    if faltas > 0:
                        dias_laborados = 15 - faltas
                        ws.cell(row=row, column=neto_quincenal_col).value = f"={bruto_mensual_neto_letter}{row}/2/15*{dias_laborados}"
                    else:
                        ws.cell(row=row, column=neto_quincenal_col).value = f"={bruto_mensual_neto_letter}{row}/2"
                        
                if descuento_col and desc > 0:
                    ws.cell(row=row, column=descuento_col).value = desc
                    
                if observaciones_col and obs_list:
                    ws.cell(row=row, column=observaciones_col).value = "; ".join(obs_list)
                    
                if puntualidad_col and sdi_letter:
                    if (punt == "NO" or retardos >= 3) and not forzar_punt:
                        ws.cell(row=row, column=puntualidad_col).value = 0.0
                    else:
                        ws.cell(row=row, column=puntualidad_col).value = f"={sdi_letter}{row}*0.1*{dias_mes_cell}"
                    
                if asistencia_col and sdi_letter:
                    if (asist == "NO" or faltas > 0) and not forzar_asist:
                        ws.cell(row=row, column=asistencia_col).value = 0.0
                    else:
                        ws.cell(row=row, column=asistencia_col).value = f"={sdi_letter}{row}*0.1*{dias_mes_cell}"
                        
                # Overridden/proportional vales
                if vales_col:
                    if ajuste_vales is not None:
                        ws.cell(row=row, column=vales_col).value = ajuste_vales
                    else:
                        effective_faltas = 0 if forzar_val else faltas
                        if effective_faltas > 0:
                            ws.cell(row=row, column=vales_col).value = f"={uma_cell}*({vales_pct_cell}/100)*{dias_mes_cell}/15*(15-{effective_faltas})"
                        else:
                            ws.cell(row=row, column=vales_col).value = f"={uma_cell}*({vales_pct_cell}/100)*{dias_mes_cell}"
                            
                # Overridden fondo_ahorro
                if fa_col:
                    if ajuste_fa is not None:
                        ws.cell(row=row, column=fa_col).value = ajuste_fa
                    else:
                        ws.cell(row=row, column=fa_col).value = f'=IF({get_field_letter(schema, "fondo_ahorro_activo")}{row}="SI", MIN({get_field_letter(schema, "sueldo_nominal")}{row}*({fa_pct_cell}/100), 1.3*{uma_cell}*{dias_mes_cell}), 0)'
                    
                # Write dynamic deductions back to Hoja1
                for col in schema.get("columns", []):
                    if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                        val = emp_agg.get(col.get("field"), 0.0)
                        if val > 0:
                            ws.cell(row=row, column=col.get("index")).value = val
                    
        row += 1

def upgrade_model_name(model_name):
    if not model_name:
        return "gemini-2.5-flash"
    m = model_name.strip()
    if m == "gemini-2.0-flash":
        return "gemini-2.5-flash"
    if m == "gemini-2.0-pro":
        return "gemini-2.5-pro"
    if m == "google/gemini-2.0-flash":
        return "google/gemini-2.5-flash"
    if m == "google/gemini-2.0-pro":
        return "google/gemini-2.5-pro"
    return m

def load_schema():
    default_schema = {"columns": [], "uma_cell": "S3", "period": "16 al 30 Abr 2026", "ai_provider": "google", "ai_model": "gemini-2.5-flash", "pending_clarifications": [], "payroll_rules": ""}
    
    path_to_load = SCHEMA_PATH
    if not os.path.exists(path_to_load):
        path_to_load = os.path.join(BASE_DIR, "schema.json")
        
    if not os.path.exists(path_to_load):
        docx_local = os.path.join(BASE_DIR, "CALCULO DE LA PRENOMINA.docx")
        default_schema["payroll_rules"] = extract_default_payroll_rules(docx_local)
        return default_schema
    try:
        with open(path_to_load, "r", encoding="utf-8") as f:
            data = json.load(f)
            
            # Auto-upgrade deprecated Gemini 2.0 models to 2.5
            original_model = data.get("ai_model", "")
            upgraded_model = upgrade_model_name(original_model)
            if original_model != upgraded_model:
                data["ai_model"] = upgraded_model
                # Try to save updated schema back to persist it
                try:
                    with open(SCHEMA_PATH, "w", encoding="utf-8") as out_f:
                        json.dump(data, out_f, ensure_ascii=False, indent=2)
                except Exception as save_err:
                    print("Error persisting upgraded schema:", save_err)

            if "payroll_rules" not in data or not data["payroll_rules"]:
                docx_local = os.path.join(BASE_DIR, "CALCULO DE LA PRENOMINA.docx")
                data["payroll_rules"] = extract_default_payroll_rules(docx_local)
            return data
    except Exception as e:
        print("Error loading schema.json:", e)
        return default_schema

def save_schema(schema):
    try:
        with open(SCHEMA_PATH, "w", encoding="utf-8") as f:
            json.dump(schema, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("Error saving schema.json:", e)

def is_path_writable(path):
    if not path:
        return False
    try:
        dir_path = os.path.dirname(os.path.abspath(path))
        if not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)
        # Test write capability
        test_file = os.path.join(dir_path, f".write_test_{os.path.basename(path)}")
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
        return True
    except Exception:
        return False

def get_excel_path():
    try:
        schema = load_schema()
        db_path = schema.get("db_path", "")
        if db_path:
            if os.path.isabs(db_path):
                if is_path_writable(db_path):
                    return db_path
                else:
                    fallback_path = os.path.abspath(os.path.join(CONFIG_DIR, os.path.basename(db_path)))
                    print(f"Database path '{db_path}' is not writable. Falling back to '{fallback_path}'")
                    return fallback_path
            # Check relative to CONFIG_DIR or BASE_DIR
            opt1 = os.path.abspath(os.path.join(CONFIG_DIR, db_path))
            if is_path_writable(opt1):
                return opt1
            if CONFIG_DIR != BASE_DIR:
                if is_path_writable(opt1):
                    return opt1
            opt2 = os.path.abspath(os.path.join(BASE_DIR, db_path))
            if is_path_writable(opt2):
                return opt2
    except Exception as e:
        print("Error getting Excel path from schema:", e)
    
    filename = "Nomina ciega.xlsx"
    if getattr(sys, 'frozen', False) or CONFIG_DIR != BASE_DIR:
        fallback_path = os.path.abspath(os.path.join(CONFIG_DIR, filename))
    else:
        fallback_path = os.path.abspath(os.path.join(BASE_DIR, filename))
        
    if not is_path_writable(fallback_path):
        fallback_path = os.path.abspath(os.path.join("/tmp", filename))
        
    return fallback_path


def find_headers_row(ws):
    # Scan the first 15 rows to find the row containing header cells
    for r in range(1, 16):
        # We check the first 10 columns
        for c in range(1, min(ws.max_column + 1, 11)):
            val = ws.cell(row=r, column=c).value
            if val and any(x in str(val).upper() for x in ["NOMBRE COMPLETO", "COD.", "NO."]):
                return r
    return 5 # Fallback to 5 if not found

def copy_template_if_needed(db_path):
    if os.path.exists(db_path):
        return
    
    # If it is an XLSX file, copy bundled excel template
    if db_path.lower().endswith(".xlsx"):
        bundled_excel = os.path.join(STATIC_DIR, "Nomina_Plantilla.xlsx")
        if not os.path.exists(bundled_excel):
            bundled_excel = os.path.join(BASE_DIR, "Nomina_Plantilla.xlsx")
            
        if os.path.exists(bundled_excel):
            print(f"Excel database file missing. Copying template to: {db_path}")
            try:
                os.makedirs(os.path.dirname(db_path), exist_ok=True)
                shutil.copyfile(bundled_excel, db_path)
            except Exception as e:
                print("Error copying template Excel:", e)
    elif db_path.lower().endswith(".csv"):
        # If it is CSV, load bundled excel and save it as CSV
        bundled_excel = os.path.join(STATIC_DIR, "Nomina_Plantilla.xlsx")
        if not os.path.exists(bundled_excel):
            bundled_excel = os.path.join(BASE_DIR, "Nomina_Plantilla.xlsx")
            
        if os.path.exists(bundled_excel):
            print(f"CSV Database file missing. Converting template to CSV: {db_path}")
            try:
                os.makedirs(os.path.dirname(db_path), exist_ok=True)
                wb = openpyxl.load_workbook(bundled_excel, data_only=True)
                ws = wb.active
                import csv
                with open(db_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    for r in range(1, ws.max_row + 1):
                        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                        writer.writerow(row_vals)
                wb.close()
            except Exception as e:
                print("Error creating CSV template:", e)

def load_workbook_agnostic(path, data_only=False):
    if path.lower().endswith(".csv"):
        import csv
        wb = openpyxl.Workbook()
        ws = wb.active
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.reader(f)
                    for r_idx, row in enumerate(reader, start=1):
                        for c_idx, val in enumerate(row, start=1):
                            if val is not None:
                                val_str = str(val).strip()
                                if val_str.startswith("="):
                                    ws.cell(row=r_idx, column=c_idx).value = val_str
                                else:
                                    try:
                                        if "." in val_str:
                                            ws.cell(row=r_idx, column=c_idx).value = float(val_str)
                                        else:
                                            ws.cell(row=r_idx, column=c_idx).value = int(val_str)
                                    except ValueError:
                                        ws.cell(row=r_idx, column=c_idx).value = val
                            else:
                                ws.cell(row=r_idx, column=c_idx).value = None
            except Exception as e:
                print(f"Error loading CSV to workbook: {e}")
        return wb
    else:
        return openpyxl.load_workbook(path, data_only=data_only)

def save_workbook_agnostic(wb, path):
    tmp_path = path + ".tmp"
    if path.lower().endswith(".csv"):
        import csv
        ws = wb.active
        try:
            with open(tmp_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                max_r = ws.max_row
                max_c = ws.max_column
                # Find last non-empty row index to avoid endless empty trailing rows
                last_non_empty = 0
                for r in range(1, max_r + 1):
                    row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
                    if any(x is not None and str(x).strip() != "" for x in row_vals):
                        last_non_empty = r
                
                for r in range(1, last_non_empty + 1):
                    row_vals = []
                    for c in range(1, max_c + 1):
                        row_vals.append(ws.cell(row=r, column=c).value)
                    writer.writerow(row_vals)
        except Exception as e:
            print(f"Error saving CSV: {e}")
            if os.path.exists(tmp_path): os.remove(tmp_path)
            raise e
    else:
        try:
            wb.save(tmp_path)
        except Exception as e:
            print(f"Error saving Excel: {e}")
            if os.path.exists(tmp_path): os.remove(tmp_path)
            raise e
            
    # Atomic rename (replace existing file)
    try:
        os.replace(tmp_path, path)
    except OSError as e:
        print(f"Error replacing file: {e}")
        raise Exception("No se pudo guardar. ¿Tienes el archivo abierto en Excel?")

def select_file_via_dialog():
    # 1. Try AppleScript on macOS first (thread-safe, out-of-process)
    if sys.platform == "darwin":
        import subprocess
        try:
            # Command System Events to activate (bringing its dialog window to the front)
            cmd = "osascript -e 'tell application \"System Events\"' -e 'activate' -e 'POSIX path of (choose file with prompt \"Seleccione el archivo de PrenÃ³mina\")' -e 'end tell'"
            proc = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if proc.returncode == 0:
                path = proc.stdout.strip()
                if path:
                    return path
            else:
                print("osascript select_file_via_dialog failed:")
                print("Exit code:", proc.returncode)
                print("Stdout:", proc.stdout)
                print("Stderr:", proc.stderr)
            return None
        except Exception as e:
            print("Failed to open dialog via osascript:", e)

    # 2. Try PowerShell on Windows (thread-safe, out-of-process)
    if sys.platform == "win32":
        import subprocess
        try:
            cmd = (
                "powershell -Command \""
                "Add-Type -AssemblyName System.Windows.Forms; "
                "$f = New-Object System.Windows.Forms.OpenFileDialog; "
                "$f.Filter = 'NÃ³mina Files (*.xlsx;*.csv)|*.xlsx;*.csv'; "
                "if ($f.ShowDialog() -eq 'OK') { $f.FileName }\""
            )
            proc = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if proc.returncode == 0:
                path = proc.stdout.strip()
                if path:
                    return path
            return None
        except Exception as e:
            print("Failed to open dialog via powershell:", e)

    # 3. Fallback to pywebview create_file_dialog (only if available)
    try:
        import webview
        if hasattr(webview, "windows") and webview.windows:
            win = webview.windows[0]
            res = win.create_file_dialog(
                dialogue_type=webview.OPEN_DIALOG,
                file_types=('Archivos de NÃ³mina (*.xlsx;*.csv)', 'Excel (*.xlsx)', 'CSV (*.csv)', 'Todos (*.*)')
            )
            if res:
                return res[0] if isinstance(res, (list, tuple)) else res
            return None
    except Exception as e:
        print("Failed to open dialog via pywebview:", e)
    return None

def select_rules_file_via_dialog():
    # 1. Try AppleScript on macOS first (thread-safe, out-of-process)
    if sys.platform == "darwin":
        import subprocess
        try:
            # Command System Events to activate (bringing its dialog window to the front)
            cmd = "osascript -e 'tell application \"System Events\"' -e 'activate' -e 'POSIX path of (choose file with prompt \"Seleccione el archivo de Reglas de NÃ³mina\")' -e 'end tell'"
            proc = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if proc.returncode == 0:
                path = proc.stdout.strip()
                if path:
                    return path
            else:
                print("osascript select_rules_file_via_dialog failed:")
                print("Exit code:", proc.returncode)
                print("Stdout:", proc.stdout)
                print("Stderr:", proc.stderr)
            return None
        except Exception as e:
            print("Failed to open rules dialog via osascript:", e)

    # 2. Try PowerShell on Windows (thread-safe, out-of-process)
    if sys.platform == "win32":
        import subprocess
        try:
            cmd = (
                "powershell -Command \""
                "Add-Type -AssemblyName System.Windows.Forms; "
                "$f = New-Object System.Windows.Forms.OpenFileDialog; "
                "$f.Filter = 'Rules Files (*.txt;*.md;*.docx)|*.txt;*.md;*.docx'; "
                "if ($f.ShowDialog() -eq 'OK') { $f.FileName }\""
            )
            proc = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if proc.returncode == 0:
                path = proc.stdout.strip()
                if path:
                    return path
            return None
        except Exception as e:
            print("Failed to open rules dialog via powershell:", e)

    # 3. Fallback to pywebview
    try:
        import webview
        if hasattr(webview, "windows") and webview.windows:
            win = webview.windows[0]
            res = win.create_file_dialog(
                dialogue_type=webview.OPEN_DIALOG,
                file_types=('Archivos de Reglas (*.txt;*.md;*.docx)', 'Documento de Word (*.docx)', 'Texto Plano (*.txt;*.md)', 'Todos (*.*)')
            )
            if res:
                return res[0] if isinstance(res, (list, tuple)) else res
            return None
    except Exception as e:
        print("Failed to open rules dialog via pywebview:", e)
    return None

# Extract database template if missing at start
copy_template_if_needed(get_excel_path())

def get_field_index(schema, field_name):
    for col in schema["columns"]:
        if col["field"] == field_name:
            return col["index"]
    return None

def get_field_letter(schema, field_name):
    for col in schema["columns"]:
        if col["field"] == field_name:
            return col["letter"]
    return ""

def get_ai_api_key(provider="google"):
    """Read AI API key from: 1) secrets.json, 2) env var. Never from schema.json."""
    secrets = load_secrets()
    key = ""
    if provider == "openrouter":
        if "openrouter_api_key" in secrets and secrets["openrouter_api_key"].strip():
            key = secrets["openrouter_api_key"].strip()
        else:
            env_key = os.environ.get("OPENROUTER_API_KEY", "").strip() or os.environ.get("OPEN_ROUTER_API_KEY", "").strip()
            if env_key:
                key = env_key
            else:
                # Check generic slots only if they look like OpenRouter keys (usually start with sk-or-)
                generic_sec = secrets.get("ai_api_key", "").strip()
                if generic_sec.startswith("sk-or-"):
                    key = generic_sec
                else:
                    generic_env = os.environ.get("AI_API_KEY", "").strip()
                    if generic_env.startswith("sk-or-"):
                        key = generic_env
        
    else:  # Google Gemini
        if "google_api_key" in secrets and secrets["google_api_key"].strip():
            key = secrets["google_api_key"].strip()
        else:
            env_key = os.environ.get("GEMINI_API_KEY", "").strip() or os.environ.get("GOOGLE_API_KEY", "").strip()
            if env_key:
                key = env_key
            else:
                # Check generic slots only if they do NOT look like OpenRouter keys
                generic_sec = secrets.get("ai_api_key", "").strip()
                if generic_sec and not generic_sec.startswith("sk-or-"):
                    key = generic_sec
                else:
                    generic_env = os.environ.get("AI_API_KEY", "").strip()
                    if generic_env and not generic_env.startswith("sk-or-"):
                        key = generic_env

    return key.strip().strip("'\"")



def get_ai_config(schema):
    """Return full AI configuration: provider, model, api_key."""
    provider = schema.get("ai_provider", "google").strip().lower()
    if provider == "none":
        return {"provider": "none", "model": "", "api_key": ""}
    model = upgrade_model_name(schema.get("ai_model", "gemini-2.5-flash"))
    api_key = get_ai_api_key(provider)
    return {"provider": provider, "model": model, "api_key": api_key}


def call_ai_api_simple(prompt, schema, response_json=True):
    """Simple AI call (e.g. for schema healing). Returns parsed JSON or raw text."""
    config = get_ai_config(schema)
    if not config["api_key"]:
        return None
    
    provider = config["provider"]
    model = config["model"]
    api_key = config["api_key"]
    
    try:
        if provider == "openrouter":
            url = "https://openrouter.ai/api/v1/chat/completions"
            req_data = {
                "model": model,
                "messages": [{"role": "user", "content": prompt}]
            }
            if response_json:
                req_data["response_format"] = {"type": "json_object"}
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://rhm-prenomina.up.railway.app",
                "X-Title": "RHM Prenomina"
            }
            data = json.dumps(req_data).encode("utf-8")
            req = urllib.request.Request(url, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=15) as response:
                res_body = response.read().decode("utf-8")
                res_json = json.loads(res_body)
                text = res_json["choices"][0]["message"]["content"]
                return json.loads(text) if response_json else text
        else:
            # Google AI (default)
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
            req_data = {
                "contents": [{"parts": [{"text": prompt}]}]
            }
            if response_json:
                req_data["generationConfig"] = {"responseMimeType": "application/json"}
            headers = {"Content-Type": "application/json"}
            data = json.dumps(req_data).encode("utf-8")
            req = urllib.request.Request(url, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=15) as response:
                res_body = response.read().decode("utf-8")
                res_json = json.loads(res_body)
                text = res_json["candidates"][0]["content"]["parts"][0]["text"]
                return json.loads(text) if response_json else text
    except Exception as e:
        print(f"Error calling AI API ({provider}/{model}):", e)
        return None

def call_ai_api_chat(system_prompt, contents_google_fmt, schema):
    """Chat-style AI call with system prompt and conversation history.
    contents_google_fmt is in Google format: [{"role": ..., "parts": [{"text": ...}]}]
    Returns the raw text response or None on failure."""
    config = get_ai_config(schema)
    if not config["api_key"]:
        return None
    
    provider = config["provider"]
    model = config["model"]
    api_key = config["api_key"]
    
    if provider == "openrouter":
        # Convert Google format to OpenAI/OpenRouter format
        messages = [{"role": "system", "content": system_prompt}]
        for entry in contents_google_fmt:
            role = entry.get("role", "user")
            text = entry["parts"][0]["text"] if entry.get("parts") else ""
            messages.append({
                "role": "user" if role == "user" else "assistant",
                "content": text
            })
        
        url = "https://openrouter.ai/api/v1/chat/completions"
        req_data = {
            "model": model,
            "messages": messages
        }
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
            "HTTP-Referer": "https://rhm-prenomina.up.railway.app",
            "X-Title": "RHM Prenomina"
        }
        data = json.dumps(req_data).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        with urllib.request.urlopen(req, timeout=60) as response:
            res_body = response.read().decode("utf-8")
            res_json = json.loads(res_body)
            return res_json["choices"][0]["message"]["content"]
    else:
        # Google AI (default)
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
        req_data = {
            "contents": contents_google_fmt,
            "systemInstruction": {
                "parts": [{"text": system_prompt}]
            }
        }
        headers = {"Content-Type": "application/json"}
        data = json.dumps(req_data).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        with urllib.request.urlopen(req, timeout=60) as response:
            res_body = response.read().decode("utf-8")
            res_json = json.loads(res_body)
            return res_json["candidates"][0]["content"]["parts"][0]["text"]

def heal_schema_locally(current_headers, old_schema):
    """
    Deterministically aligns old schema columns with new Excel headers,
    and automatically registers any new columns using local rules.
    """
    import re
    import openpyxl.utils

    updated_schema = dict(old_schema)
    old_columns = old_schema.get("columns", [])
    
    # Map old columns by header (case-insensitive, stripped)
    old_by_header = {}
    for col in old_columns:
        h = col.get("header")
        if h:
            old_by_header[str(h).strip().upper()] = col

    new_columns = []
    seen_fields = set()

    def clean_field_name(header_str, index):
        # Convert header to a valid snake_case field name
        s = header_str.lower().strip()
        s = re.sub(r"[^\w\s-]", "", s)
        s = re.sub(r"[\s-]+", "_", s)
        s = s.strip("_")
        if not s:
            return f"col_{index}"
        # Avoid duplicate fields
        orig = s
        counter = 1
        while s in seen_fields:
            s = f"{orig}_{counter}"
            counter += 1
        return s

    for idx, h_raw in enumerate(current_headers):
        h_str = str(h_raw).strip() if h_raw is not None else ""
        
        h_upper = h_str.upper()
        col_idx = idx + 1
        letter = openpyxl.utils.get_column_letter(col_idx)

        # Check if this header already exists in old schema
        if h_str and h_upper in old_by_header:
            # Re-align existing column
            col = dict(old_by_header[h_upper])
            col["index"] = col_idx
            col["letter"] = letter
            new_columns.append(col)
            seen_fields.add(col["field"])
        else:
            # This is a NEW or empty column!
            # If it's completely empty, check if we can reuse an existing column from old_schema that has an empty header
            old_col_at_idx = None
            for col in old_columns:
                if col.get("index") == col_idx and (col.get("header") is None or str(col.get("header")).strip() == ""):
                    old_col_at_idx = col
                    break
            
            if old_col_at_idx:
                col = dict(old_col_at_idx)
                col["index"] = col_idx
                col["letter"] = letter
                new_columns.append(col)
                seen_fields.add(col["field"])
            else:
                field_name = clean_field_name(h_str, col_idx)
                seen_fields.add(field_name)

                # Determine type and category based on header text
                category = "metadata"
                col_type = "string"
                editable = True
                incidence_editable = False

                h_lower = h_str.lower()
                # Simple heuristic matching
                if h_str:
                    if any(x in h_lower for x in ["fecha", "ingreso", "baja"]):
                        col_type = "date"
                    elif any(x in h_lower for x in ["cuenta", "activo", "s/n", "si/no"]):
                        col_type = "boolean"
                    elif any(x in h_lower for x in ["descuento", "deuda", "prestamo", "abono"]):
                        col_type = "float"
                        category = "deduction"
                        incidence_editable = True
                    elif any(x in h_lower for x in ["total", "suma", "sdi", "puntualidad", "asistencia", "nominal", "integracion", "acumulado", "bruto"]):
                        col_type = "float"
                        category = "calculated"
                        editable = False
                    elif any(x in h_lower for x in ["diario", "sueldo"]):
                        col_type = "float"
                        category = "nominal_imss"
                    elif any(x in h_lower for x in ["gasolina", "combustible", "bono", "comision", "efectivo", "facturado", "asimilados", "socio"]):
                        col_type = "float"
                        category = "others"
                    elif any(x in h_lower for x in ["monto", "pago", "cantidad", "pesos", "$", "%"]):
                        col_type = "float"
                        category = "others"

                new_col = {
                    "index": col_idx,
                    "letter": letter,
                    "header": h_str if h_str else None,
                    "field": field_name,
                    "type": col_type,
                    "category": category,
                    "label": h_str if h_str else f"Columna {letter}",
                    "editable": editable if h_str else False,
                    "incidence_editable": incidence_editable
                }
                new_columns.append(new_col)
                print(f"Locally detected and added column: {h_str or 'VACIA'} (field: {field_name}, category: {category})")

                # If we defaulted to metadata but it could be a monetary field, ask the user
                if h_str and category == "metadata" and col_type == "string":
                    if "pending_clarifications" not in updated_schema:
                        updated_schema["pending_clarifications"] = []
                    
                    # Avoid duplicates
                    if not any(q.get("field") == field_name for q in updated_schema["pending_clarifications"]):
                        updated_schema["pending_clarifications"].append({
                            "field": field_name,
                            "question": f"He detectado una nueva columna '{h_str}'. Â¿CÃ³mo debe procesarse en la prenÃ³mina?",
                            "options": [
                                "Es una DeducciÃ³n (Descuento)",
                                "Es una PercepciÃ³n Adicional",
                                "Solo es texto informativo"
                            ]
                        })

    # Sort columns by index
    updated_schema["columns"] = sorted(new_columns, key=lambda x: x["index"])
    return updated_schema

def heal_schema_with_ai(current_headers, old_schema):
    provider = old_schema.get("ai_provider", "google").strip().lower()
    api_key = get_ai_api_key(provider)
    if not api_key:
        print("AI API Key is missing. Performing local deterministic schema healing.")
        updated_schema = heal_schema_locally(current_headers, old_schema)
        # Create a pending clarification to guide the user to configure key for advanced AI healing
        updated_schema["pending_clarifications"] = [{
            "field": "ai_api_key",
            "question": "Se han detectado cambios en las cabeceras de Excel, pero no hay una Clave de API de IA configurada. Se han alineado las columnas de forma local, pero puedes introducir tu clave en ConfiguraciÃ³n para usar la clasificaciÃ³n inteligente.",
            "options": ["Reintentar con clave configurada", "Omitir por ahora"]
        }]
        return updated_schema

    old_columns_summary = []
    for col in old_schema["columns"]:
        old_columns_summary.append({
            "index": col["index"],
            "letter": col["letter"],
            "header": col["header"],
            "field": col["field"],
            "category": col["category"]
        })

    prompt = f"""
    You are an expert AI data engineer specialized in payroll system integrations.
    The user has modified their Excel payroll spreadsheet. I will give you:
    1. The old schema configuration mapping Excel columns to database fields.
    2. The new list of headers read from row 5 of the Excel sheet.

    Old columns mapping:
    {json.dumps(old_columns_summary, ensure_ascii=False, indent=2)}

    New Excel headers (1-based index):
    {json.dumps([{ 'index': idx + 1, 'header': header } for idx, header in enumerate(current_headers) if header is not None], ensure_ascii=False, indent=2)}

    Your task:
    1. Align the columns index and letters to match the new headers. If headers shifted, update their index and letter accordingly.
    2. Identify any NEW headers that were not in the old columns mapping.
    3. Categorize new headers as one of the following:
       - 'metadata': General info (text, dates, or boolean fields like flags)
       - 'nominal_imss': Core IMSS daily salary or related fields (user editable float)
       - 'others': Other monthly payment schemas (commissions, bonuses, petrol, etc. - user editable float)
       - 'deduction': Incidences / discounts to apply (e.g. loans, debts - user editable float)
       - 'calculated': Column calculated via formulas (e.g. SDI, Sueldo Nominal, sums)
    4. For any new header, assign it a unique snake_case 'field' identifier (e.g. 'bono_asistencia', 'prestamo_personal').
    5. Set 'type': 'float' for numeric payment fields, 'date' for date strings, 'boolean' for SI/NO flags, 'string' for general text.
    6. Set 'label': A short clean UI label in Spanish (e.g. 'Bono de Asistencia ($)').
    7. Set 'editable': true for metadata, nominal_imss, others.
    8. Set 'incidence_editable': true for deductions.
    9. If a new column is a new deduction or calculation variable and you are not 100% sure if it is a flat deduction or needs division (like a loan split over months), do NOT guess. Instead, add a clarification question to the 'pending_clarifications' list in the format:
       {{
         "field": "the_new_field_identifier",
         "question": "A short clear question in Spanish asking the accountant how to calculate/apply this new column.",
         "options": ["Option A text in Spanish", "Option B text in Spanish"]
       }}
    10. Return a JSON object with:
        {{
          "columns": [ ... ],
          "pending_clarifications": [ ... ]
        }}
        The 'columns' array must contain the complete set of columns (both old/updated and new ones), sorted by index.

    Respond strictly in JSON format.
    """
    res = call_ai_api_simple(prompt, old_schema, response_json=True)
    if res and "columns" in res:
        updated_schema = dict(old_schema)
        old_fields = {col["field"]: col for col in old_schema["columns"]}
        merged_cols = []
        for col in res["columns"]:
            f_name = col["field"]
            letter = openpyxl.utils.get_column_letter(col["index"])
            col["letter"] = letter
            if f_name in old_fields:
                col["editable"] = old_fields[f_name].get("editable", col.get("editable", True))
                col["incidence_editable"] = old_fields[f_name].get("incidence_editable", col.get("incidence_editable", False))
                if "formula_template" in old_fields[f_name]:
                    col["formula_template"] = old_fields[f_name]["formula_template"]
            merged_cols.append(col)
        updated_schema["columns"] = sorted(merged_cols, key=lambda x: x["index"])
        updated_schema["pending_clarifications"] = res.get("pending_clarifications", [])
        return updated_schema
    else:
        print("AI Schema healing failed or timed out. Falling back to local deterministic schema healing.")
        return heal_schema_locally(current_headers, old_schema)

def ensure_vacation_columns_in_excel(excel_path, schema):
    if not os.path.exists(excel_path):
        return
    try:
        wb = load_workbook_agnostic(excel_path, data_only=False)
        ws = wb.active
        headers_row = find_headers_row(ws)
        
        # Read the headers of this row
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=headers_row, column=c).value
            headers.append(val)
            
        while headers and (headers[-1] is None or str(headers[-1]).strip() == ""):
            headers.pop()
            
        headers_upper = [str(h).upper().strip() for h in headers if h is not None]
        
        needed_columns = [
            ("Vacaciones Totales / Derecho", "vacaciones_totales", 38, "AL"),
            ("Vacaciones Tomadas", "vacaciones_tomadas", 39, "AM"),
            ("Vacaciones Restantes / Disponibles", "vacaciones_restantes", 40, "AN")
        ]
        
        modified = False
        for name, field, target_idx, letter in needed_columns:
            found = False
            for h in headers:
                if h and str(h).upper().strip() in [
                    name.upper(), 
                    "VACACIONES TOTALES", 
                    "VACACIONES DERECHO", 
                    "VACACIONES TOMADAS", 
                    "VACACIONES RESTANTES", 
                    "VACACIONES DISPONIBLES",
                    "VACACIONES TOTALES / DERECHO",
                    "VACACIONES RESTANTES / DISPONIBLES"
                ]:
                    found = True
                    break
            if not found:
                print(f"Adding column '{name}' at index {target_idx} ({letter}) in Excel file...")
                ws.cell(row=headers_row, column=target_idx).value = name
                modified = True
                
        if modified:
            save_workbook_agnostic(wb, excel_path)
            
        wb.close()
    except Exception as e:
        print("Error ensuring vacation columns in Excel:", e)

def check_and_heal_schema():
    schema = load_schema()
    excel_path = get_excel_path()
    ensure_vacation_columns_in_excel(excel_path, schema)
    if not os.path.exists(excel_path):
        return schema
    try:
        wb = load_workbook_agnostic(excel_path, data_only=True)
        ws = wb.active
        headers_row = find_headers_row(ws)
        current_headers = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=headers_row, column=col_idx).value
            current_headers.append(val)
        wb.close()

        while current_headers and (current_headers[-1] is None or str(current_headers[-1]).strip() == ""):
            current_headers.pop()

        schema_headers = [col.get("header") for col in schema["columns"]]
        while schema_headers and (schema_headers[-1] is None or str(schema_headers[-1]).strip() == ""):
            schema_headers.pop()

        mismatch = False
        if len(current_headers) != len(schema_headers):
            mismatch = True
        else:
            for ch, sh in zip(current_headers, schema_headers):
                ch_str = str(ch).strip() if ch is not None else ""
                sh_str = str(sh).strip() if sh is not None else ""
                if ch_str != sh_str:
                    mismatch = True
                    break

        if mismatch:
            print("Excel headers mismatch detected! Running self-healing schema wrapper...")
            new_schema = heal_schema_with_ai(current_headers, schema)
            save_schema(new_schema)
            
            # Rewrite formulas in Hoja1 to align with newly healed column letters
            try:
                wb_re = load_workbook_agnostic(excel_path, data_only=False)
                ws_re = wb_re.active
                nombre_col = get_field_index(new_schema, "nombre")
                id_col = get_field_index(new_schema, "id")
                headers_row = find_headers_row(ws_re)
                row = headers_row + 1
                while True:
                    nombre_val = ws_re.cell(row=row, column=nombre_col).value
                    cod_val = ws_re.cell(row=row, column=id_col).value
                    if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                        break
                    if nombre_val is None and cod_val is None:
                        break
                    if nombre_val:
                        inject_formulas_dynamically(ws_re, row, new_schema)
                    row += 1
                save_workbook_agnostic(wb_re, excel_path)
                wb_re.close()
                print("Excel formulas re-injected successfully after schema self-healing.")
            except Exception as reinj_err:
                print("Error re-injecting formulas after schema healing:", reinj_err)
                
            return new_schema
        return schema
    except Exception as e:
        print("Error checking schema alignment:", e)
        return schema

def inject_formulas_dynamically(ws, row, schema):
    def L(field_name):
        return get_field_letter(schema, field_name)
    
    uma_cell = ensure_absolute_cell(schema.get("uma_cell", "S3"))
    vales_pct_cell = ensure_absolute_cell(schema.get("vales_pct_cell", "P3"))
    dias_mes_cell = ensure_absolute_cell(schema.get("dias_mes_cell", "N3"))
    fa_pct_cell = ensure_absolute_cell(schema.get("fa_pct_cell", "L3"))
    
    # Build core formulas dynamically based on letter configurations
    # Subtract all active deduction columns from bruto_mensual to get bruto_mensual_neto
    ded_fields = [col["field"] for col in schema.get("columns", []) if col.get("category") == "deduction"]
    ded_sub_str = "".join([f"-{L(df)}{row}" for df in ded_fields if L(df)])
    if not ded_sub_str:
        ded_sub_str = f"-{L('descuento_adicional')}{row}"
 
    formulas = {
        "sdi": f"={L('salario_diario')}{row}*{L('factor_integracion')}{row}",
        "sueldo_nominal": f"={L('salario_diario')}{row}*{dias_mes_cell}",
        "puntualidad": f"={L('sdi')}{row}*0.1*{dias_mes_cell}",
        "asistencia": f"={L('sdi')}{row}*0.1*{dias_mes_cell}",
        "vales_despensa": f"={uma_cell}*({vales_pct_cell}/100)*{dias_mes_cell}",
        "fondo_ahorro": f'=IF({L("fondo_ahorro_activo")}{row}="SI",MIN({L("sueldo_nominal")}{row}*({fa_pct_cell}/100),1.3*{uma_cell}*{dias_mes_cell}),0)',
        "percepcion_sueldos": f"=SUM({L('sueldo_nominal')}{row}:{L('fondo_ahorro')}{row})",
        "bruto_mensual": f"=SUM({L('percepcion_sueldos')}{row}:{L('deuda_carro')}{row})",
        "bruto_quincenal": f"={L('bruto_mensual')}{row}/2",
        "bruto_mensual_neto": f"={L('bruto_mensual')}{row}{ded_sub_str}",
        "descuento_quincenal_acumulado": f"={L('bruto_quincenal')}{row}-{L('neto_quincenal')}{row}",
        "vacaciones_restantes": f"={L('vacaciones_totales')}{row}-{L('vacaciones_tomadas')}{row}"
    }
 
    # Write formulas into the spreadsheet row
    for col in schema["columns"]:
        f = col["field"]
        if f in formulas:
            ws.cell(row=row, column=col["index"]).value = formulas[f]

class APIHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=STATIC_DIR, **kwargs)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def do_GET(self):
        path_only = self.path.split("?")[0]
        
        # Static files or non-API routes are served publicly
        if not path_only.startswith("/api/"):
            super().do_GET()
            return
            
        # Authentication middleware for API routes
        session = self.get_session_user()
        if not session:
            self.send_json({"error": "No autorizado. Inicie sesiÃ³n."}, 401)
            return
            
        # Admin-only GET endpoints
        if path_only == "/api/download-excel" or path_only == "/api/users":
            if session["role"] != "admin":
                self.send_json({"error": "Prohibido. Se requieren permisos de administrador."}, 403)
                return
                
        # Endpoint routes
        with EXCEL_LOCK:
            if path_only == "/api/employees":
                self.get_employees()
            elif path_only == "/api/schema":
                self.get_schema()
            elif path_only == "/api/select-file":
                self.select_file()
            elif path_only == "/api/select-rules-file":
                self.select_rules_file()
            elif path_only == "/api/download-excel":
                self.download_excel()
            elif path_only == "/api/users":
                self.get_users()
            elif path_only == "/api/ai-status":
                self.get_ai_status()
            elif path_only == "/api/companies":
                self.get_companies()
            elif path_only == "/api/incidences":
                import urllib.parse
                parsed = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed.query)
                employee_id = params.get("id", [None])[0]
                self.get_incidences_endpoint(employee_id)
            else:
                self.send_json({"error": "Endpoint not found"}, 404)

    def do_POST(self):
        path_only = self.path.split("?")[0]
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length) if content_length > 0 else b""

        # Login is public
        if path_only == "/api/login":
            self.login_endpoint(post_data)
            return

        # Authentication middleware for other API routes
        session = self.get_session_user()
        if not session:
            self.send_json({"error": "No autorizado. Inicie sesiÃ³n."}, 401)
            return

        # Admin-only POST endpoints
        if path_only in ["/api/config", "/api/users", "/api/select-file", "/api/select-rules-file", "/api/upload-database", "/api/upload-rules", "/api/companies", "/api/schema/validate", "/api/schema/confirm"]:
            if session["role"] != "admin":
                self.send_json({"error": "Prohibido. Se requieren permisos de administrador."}, 403)
                return

        if path_only == "/api/parse-docx":
            self.parse_docx_endpoint(post_data)
            return
        elif path_only == "/api/upload-database":
            self.upload_database_endpoint(post_data)
            return
        elif path_only == "/api/upload-rules":
            self.upload_rules_endpoint(post_data)
            return

        try:
            body = json.loads(post_data.decode("utf-8")) if post_data else {}
        except Exception as e:
            self.send_json({"error": f"Invalid JSON: {e}"}, 400)
            return

        with EXCEL_LOCK:
            if path_only == "/api/collaborator":
                self.save_collaborator(body)
            elif path_only == "/api/companies":
                self.save_companies(body)
            elif path_only == "/api/incidences":
                self.save_incidences(body)
            elif path_only == "/api/period":
                self.save_period(body)
            elif path_only == "/api/config":
                self.save_config(body)
            elif path_only == "/api/schema/clarify":
                self.save_clarify(body)
            elif path_only == "/api/schema/validate":
                self.schema_validate_endpoint(body)
            elif path_only == "/api/schema/confirm":
                self.schema_confirm_endpoint(body)
            elif path_only == "/api/payroll/explain":
                self.explain_payroll(body)
            elif path_only == "/api/users":
                self.save_user_endpoint(body)
            elif path_only == "/api/logout":
                self.logout_endpoint()
            else:
                self.send_json({"error": "Endpoint not found"}, 404)

    def do_DELETE(self):
        path_only = self.path.split("?")[0]
        session = self.get_session_user()
        if not session:
            self.send_json({"error": "No autorizado. Inicie sesiÃ³n."}, 401)
            return

        with EXCEL_LOCK:
            if path_only == "/api/users":
                if session["role"] != "admin":
                    self.send_json({"error": "Prohibido. Se requieren permisos de administrador."}, 403)
                    return
                import urllib.parse
                parsed = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed.query)
                username = params.get("username", [None])[0]
                if not username:
                    self.send_json({"error": "Falta el nombre de usuario a eliminar"}, 400)
                    return
                self.delete_user_endpoint(username)
            elif path_only == "/api/companies":
                if session["role"] != "admin":
                    self.send_json({"error": "Prohibido. Se requieren permisos de administrador."}, 403)
                    return
                import urllib.parse
                parsed = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed.query)
                company_id = params.get("id", [None])[0]
                if not company_id:
                    self.send_json({"error": "Falta el ID de la empresa a eliminar"}, 400)
                    return
                self.delete_company_endpoint(company_id)
            elif path_only == "/api/incidences":
                import urllib.parse
                parsed = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed.query)
                employee_id = params.get("employee_id", [None])[0]
                date_val = params.get("date", [None])[0]
                if not employee_id or not date_val:
                    self.send_json({"error": "Falta employee_id o date para eliminar la incidencia"}, 400)
                    return
                self.delete_incidence_endpoint(employee_id, date_val)
            else:
                self.send_json({"error": "Endpoint not found"}, 404)

    def get_session_user(self):
        token = None
        auth_header = self.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
        else:
            # Fallback to query parameter 'token' for file downloads
            try:
                import urllib.parse
                parsed = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed.query)
                token_list = params.get("token")
                if token_list:
                    token = token_list[0]
            except Exception as e:
                print("Error parsing token from query string:", e)
                
        if not token:
            return None
            
        session = SESSIONS.get(token)
        if not session:
            return None
        if session["expiry"] < time.time():
            try:
                del SESSIONS[token]
            except KeyError:
                pass
            return None
        return session

    def login_endpoint(self, post_data):
        try:
            body = json.loads(post_data.decode("utf-8")) if post_data else {}
            username = body.get("username", "").strip()
            password = body.get("password", "")
            
            if not username or not password:
                self.send_json({"error": "Usuario y contraseÃ±a requeridos"}, 400)
                return
                
            users = load_users()
            user = users.get(username)
            if not user:
                self.send_json({"error": "Usuario o contraseÃ±a incorrectos"}, 401)
                return
                
            hashed = hash_password(password, user["salt"])
            if hashed != user["password"]:
                self.send_json({"error": "Usuario o contraseÃ±a incorrectos"}, 401)
                return
                
            # Create session
            token = secrets.token_hex(24)
            expiry = time.time() + 86400  # 24 hours
            SESSIONS[token] = {
                "username": username,
                "role": user["role"],
                "expiry": expiry
            }
            
            self.send_json({
                "token": token,
                "role": user["role"],
                "username": username
            })
        except Exception as e:
            self.send_json({"error": f"Error en login: {str(e)}"}, 500)

    def logout_endpoint(self):
        auth_header = self.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
            if token in SESSIONS:
                del SESSIONS[token]
        self.send_json({"success": True})

    def get_users(self):
        try:
            users = load_users()
            public_users = []
            for u in users.values():
                public_users.append({
                    "username": u["username"],
                    "role": u["role"]
                })
            self.send_json(public_users)
        except Exception as e:
            self.send_json({"error": f"Error obteniendo usuarios: {e}"}, 500)

    def save_user_endpoint(self, body):
        try:
            username = body.get("username", "").strip()
            password = body.get("password", "")
            role = body.get("role", "capturista")
            
            if not username:
                self.send_json({"error": "El nombre de usuario es requerido"}, 400)
                return
                
            if role not in ["admin", "capturista"]:
                self.send_json({"error": "Rol invÃ¡lido"}, 400)
                return
                
            users = load_users()
            
            if username in users:
                if password:
                    salt = secrets.token_hex(8)
                    users[username]["salt"] = salt
                    users[username]["password"] = hash_password(password, salt)
                users[username]["role"] = role
            else:
                if not password:
                    self.send_json({"error": "La contraseÃ±a es requerida para nuevos usuarios"}, 400)
                    return
                salt = secrets.token_hex(8)
                users[username] = {
                    "username": username,
                    "role": role,
                    "salt": salt,
                    "password": hash_password(password, salt)
                }
                
            save_users(users)
            self.send_json({"success": True})
        except Exception as e:
            self.send_json({"error": f"Error guardando usuario: {e}"}, 500)

    def delete_user_endpoint(self, username):
        try:
            session = self.get_session_user()
            if session and session["username"] == username:
                self.send_json({"error": "No puedes eliminar tu propio usuario activo"}, 400)
                return
                
            users = load_users()
            if username not in users:
                self.send_json({"error": "El usuario no existe"}, 404)
                return
                
            if users[username]["role"] == "admin":
                admins = [u for u in users.values() if u["role"] == "admin"]
                if len(admins) <= 1:
                    self.send_json({"error": "Debe existir al menos un usuario administrador en el sistema"}, 400)
                    return
                    
            del users[username]
            save_users(users)
            self.send_json({"success": True})
        except Exception as e:
            self.send_json({"error": f"Error eliminando usuario: {e}"}, 500)

    def get_companies(self):
        try:
            companies = load_companies()
            self.send_json(companies)
        except Exception as e:
            self.send_json({"error": f"Error al obtener catálogo de empresas: {e}"}, 500)

    def save_companies(self, body):
        try:
            company_id = body.get("id")
            nombre = body.get("nombre", "").strip()
            razon_social = body.get("razon_social", "").strip()
            regimen = body.get("regimen", "").strip()
            try:
                prima_riesgo = float(str(body.get("prima_riesgo", 0.0)).replace("%", "").strip())
            except ValueError:
                prima_riesgo = 0.0
            
            if not nombre:
                self.send_json({"error": "El nombre corto de la empresa es requerido"}, 400)
                return
                
            companies = load_companies()
            
            if company_id:
                found = False
                for c in companies:
                    if str(c["id"]) == str(company_id):
                        c["nombre"] = nombre
                        c["razon_social"] = razon_social
                        c["regimen"] = regimen
                        c["prima_riesgo"] = prima_riesgo
                        found = True
                        break
                if not found:
                    self.send_json({"error": f"Empresa con ID {company_id} no encontrada"}, 404)
                    return
            else:
                new_id = str(max([int(c["id"]) for c in companies] + [0]) + 1)
                companies.append({
                    "id": new_id,
                    "nombre": nombre,
                    "razon_social": razon_social,
                    "regimen": regimen,
                    "prima_riesgo": prima_riesgo
                })
                
            save_companies_data(companies)
            self.send_json({"success": True, "message": "Empresa guardada exitosamente"})
        except Exception as e:
            self.send_json({"error": f"Error al guardar empresa: {e}"}, 500)

    def delete_company_endpoint(self, company_id):
        try:
            companies = load_companies()
            new_companies = [c for c in companies if str(c["id"]) != str(company_id)]
            if len(new_companies) == len(companies):
                self.send_json({"error": f"Empresa con ID {company_id} no encontrada"}, 404)
                return
            
            save_companies_data(new_companies)
            self.send_json({"success": True, "message": "Empresa eliminada exitosamente"})
        except Exception as e:
            self.send_json({"error": f"Error al eliminar empresa: {e}"}, 500)

    def get_ai_status(self):
        try:
            schema = load_schema()
            provider = schema.get("ai_provider", "google").strip().lower()
            api_key = get_ai_api_key(provider)
            self.send_json({"configured": bool(api_key)})
        except Exception as e:
            self.send_json({"error": str(e)}, 500)



    def get_incidences_endpoint(self, employee_id):
        try:
            schema = check_and_heal_schema()
            excel_path = get_excel_path()
            if not os.path.exists(excel_path):
                self.send_json([])
                return
            
            # Check for migration first using a writeable workbook
            try:
                wb_mig = load_workbook_agnostic(excel_path, data_only=False)
                if heal_incidences_sheet_if_needed(wb_mig, schema):
                    save_workbook_agnostic(wb_mig, excel_path)
                wb_mig.close()
            except Exception as mig_err:
                print(f"Error checking/migrating Incidencias in get_incidences_endpoint: {mig_err}")
                
            period_str = schema.get("period", "16 al 30 Abr 2026")
            start_date, end_date = parse_period_dates(period_str)
            
            wb = load_workbook_agnostic(excel_path, data_only=True)
            incidences_list = []
            if "Incidencias" in wb.sheetnames:
                ws_inc = wb["Incidencias"]
                for r in range(2, ws_inc.max_row + 1):
                    date_val = ws_inc.cell(row=r, column=1).value
                    r_cod = ws_inc.cell(row=r, column=2).value
                    if not date_val or not r_cod:
                        continue
                    
                    row_date = parse_date_robust(date_val)
                    if row_date and start_date <= row_date <= end_date:
                        cleaned_r_cod = clean_employee_id(r_cod)
                        if not employee_id or cleaned_r_cod == clean_employee_id(employee_id):
                            max_c = ws_inc.max_column
                            inc_item = {
                                "row_idx": r,
                                "date": row_date.strftime("%Y-%m-%d"),
                                "id": cleaned_r_cod,
                                "nombre": str(ws_inc.cell(row=r, column=3).value or ""),
                                "faltas": int(ws_inc.cell(row=r, column=4).value or 0),
                                "retardos": int(ws_inc.cell(row=r, column=5).value or 0),
                                "vacaciones": int(ws_inc.cell(row=r, column=6).value or 0),
                                "descuento_adicional": float(ws_inc.cell(row=r, column=7).value or 0.0),
                                "puntualidad": str(ws_inc.cell(row=r, column=8).value or "SI"),
                                "asistencia": str(ws_inc.cell(row=r, column=9).value or "SI"),
                                "observaciones": str(ws_inc.cell(row=r, column=10).value or ""),
                                "forzar_asistencia": str(ws_inc.cell(row=r, column=11).value or "NO") if max_c >= 11 else "NO",
                                "forzar_puntualidad": str(ws_inc.cell(row=r, column=12).value or "NO") if max_c >= 12 else "NO",
                                "forzar_vales": str(ws_inc.cell(row=r, column=13).value or "NO") if max_c >= 13 else "NO",
                                "ajuste_vales": ws_inc.cell(row=r, column=14).value if max_c >= 14 else None,
                                "ajuste_fondo_ahorro": ws_inc.cell(row=r, column=15).value if max_c >= 15 else None
                            }
                            # Clean numeric fields
                            if inc_item["ajuste_vales"] is not None and str(inc_item["ajuste_vales"]).strip() != "":
                                try: inc_item["ajuste_vales"] = float(inc_item["ajuste_vales"])
                                except ValueError: inc_item["ajuste_vales"] = None
                            else:
                                inc_item["ajuste_vales"] = None
                                
                            if inc_item["ajuste_fondo_ahorro"] is not None and str(inc_item["ajuste_fondo_ahorro"]).strip() != "":
                                try: inc_item["ajuste_fondo_ahorro"] = float(inc_item["ajuste_fondo_ahorro"])
                                except ValueError: inc_item["ajuste_fondo_ahorro"] = None
                            else:
                                inc_item["ajuste_fondo_ahorro"] = None
                                
                            # If there are dynamic fields beyond column 15, load them
                            if schema and schema.get("columns"):
                                col_idx = 16
                                for col in schema["columns"]:
                                    if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                                        if col_idx <= max_c:
                                            inc_item[col.get("field")] = float(ws_inc.cell(row=r, column=col_idx).value or 0.0)
                                            col_idx += 1
                            incidences_list.append(inc_item)
            wb.close()
            # Sort by date
            incidences_list.sort(key=lambda x: x["date"])
            self.send_json(incidences_list)
        except Exception as e:
            self.send_json({"error": f"Error loading incidences list: {e}", "details": traceback.format_exc()}, 500)

    def delete_incidence_endpoint(self, employee_id, date_val):
        try:
            schema = check_and_heal_schema()
            excel_path = get_excel_path()
            if not os.path.exists(excel_path):
                self.send_json({"error": "Database file not found"}, 500)
                return
            
            wb = load_workbook_agnostic(excel_path, data_only=False)
            ws_inc = wb["Incidencias"] if "Incidencias" in wb.sheetnames else None
            if not ws_inc:
                self.send_json({"error": "Incidencias sheet not found"}, 404)
                wb.close()
                return
            
            cleaned_emp_id = clean_employee_id(employee_id)
            target_date = parse_date_robust(date_val)
            if not target_date:
                self.send_json({"error": "Invalid date format"}, 400)
                wb.close()
                return
                
            target_date_str = target_date.strftime("%Y-%m-%d")
            
            found_row = None
            for r in range(2, ws_inc.max_row + 1):
                r_date = ws_inc.cell(row=r, column=1).value
                r_cod = ws_inc.cell(row=r, column=2).value
                r_date_parsed = parse_date_robust(r_date)
                if r_date_parsed and r_cod:
                    r_date_str = r_date_parsed.strftime("%Y-%m-%d")
                    if r_date_str == target_date_str and clean_employee_id(r_cod) == cleaned_emp_id:
                        found_row = r
                        break
            
            if found_row:
                ws_inc.delete_rows(found_row)
                recompile_active_period_incidences(wb, schema)
                save_workbook_agnostic(wb, excel_path)
                wb.close()
                self.send_json({"success": True, "message": "Incidence deleted successfully"})
            else:
                wb.close()
                self.send_json({"error": "Incidence not found"}, 404)
                
        except PermissionError:
            self.send_json({"error": f"El archivo estÃ¡ bloqueado o abierto en Excel. Por favor ciÃ©rralo e intÃ©ntalo de nuevo."}, 500)
        except Exception as e:
            self.send_json({"error": f"Error deleting incidence: {e}", "details": traceback.format_exc()}, 500)

    def parse_docx_endpoint(self, post_data):
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            import io
            
            if not post_data:
                self.send_json({"error": "Empty body"}, 400)
                return

            with zipfile.ZipFile(io.BytesIO(post_data)) as z:
                doc_xml = z.read('word/document.xml')
                root = ET.fromstring(doc_xml)
                
                paragraphs = []
                for elem in root.iter():
                    tag_name = elem.tag.split('}')[-1]
                    if tag_name == 'p':
                        p_text = []
                        for child in elem.iter():
                            child_tag = child.tag.split('}')[-1]
                            if child_tag == 't' and child.text:
                                p_text.append(child.text)
                        text_str = "".join(p_text).strip()
                        if text_str:
                            paragraphs.append(text_str)
                extracted_text = "\n\n".join(paragraphs)
            self.send_json({"text": extracted_text})
        except Exception as e:
            self.send_json({"error": f"Failed to parse docx file: {e}"}, 500)

    def upload_database_endpoint(self, post_data):
        try:
            boundary = self.headers.get_param('boundary')
            if not boundary:
                self.send_json({"error": "No boundary in Content-Type"}, 400)
                return
            
            parts = parse_multipart(post_data, boundary)
            file_part = parts.get("file")
            if not file_part:
                self.send_json({"error": "No file uploaded in form data"}, 400)
                return
                
            filename = file_part["filename"]
            content = file_part["content"]
            
            if not filename or not content:
                self.send_json({"error": "File is empty or has no name"}, 400)
                return
                
            ext = filename.split('.')[-1].lower()
            if ext not in ["xlsx", "xls", "csv"]:
                self.send_json({"error": "Formato de archivo no soportado. Debe ser .xlsx o .csv"}, 400)
                return
                
            safe_filename = "".join(c for c in filename if c.isalnum() or c in "._- ")
            dest_path = os.path.join(CONFIG_DIR, safe_filename)
            
            with open(dest_path, "wb") as f:
                f.write(content)
                
            self.send_json({"selected_path": dest_path})
        except Exception as e:
            tb = traceback.format_exc()
            print("Error in upload_database_endpoint:\n", tb)
            self.send_json({"error": f"Error al subir base de datos: {e}", "details": tb}, 500)

# HELPER FUNCTIONS FOR INTERACTIVE SCHEMA VERIFICATION AND FORMULA AUDITING
def identify_critical_headers(headers):
    nombre_idx = None
    sdi_idx = None
    
    nombre_kws = ["nombre completo", "nombre", "empleado", "colaborador", "trabajador", "nombre completo / asimilados"]
    sdi_kws = ["sdi", "salario diario integrado", "salario integrado", "integrado", "factor de integracion", "factor integracion", "s.d.i.", "factor integracion imss"]
    
    for i, h in enumerate(headers):
        if not h:
            continue
        h_clean = str(h).lower().strip()
        
        # Check exact matches first
        if h_clean == "nombre completo" or h_clean == "nombre":
            nombre_idx = i + 1
        elif h_clean == "sdi" or h_clean == "salario diario integrado" or h_clean == "factor integracion imss":
            sdi_idx = i + 1
            
    # Fallback to fuzzy keyword search
    if nombre_idx is None:
        for i, h in enumerate(headers):
            if not h:
                continue
            h_clean = str(h).lower().strip()
            if any(kw in h_clean for kw in nombre_kws):
                nombre_idx = i + 1
                break
                
    if sdi_idx is None:
        for i, h in enumerate(headers):
            if not h:
                continue
            h_clean = str(h).lower().strip()
            if any(kw in h_clean for kw in sdi_kws):
                sdi_idx = i + 1
                break
                
    return nombre_idx, sdi_idx

def validate_schema_formulas_locally(columns):
    tot_letter = "AL"
    tom_letter = "AM"
    faltas_letter = "AH"
    bruto_mensual_letter = "AB"
    bruto_mensual_2_letter = "AF"
    
    for col in columns:
        f = col.get("field")
        if f == "vacaciones_totales":
            tot_letter = col.get("letter")
        elif f == "vacaciones_tomadas":
            tom_letter = col.get("letter")
        elif f == "faltas":
            faltas_letter = col.get("letter")
        elif f == "sueldo_bruto_mensual":
            bruto_mensual_letter = col.get("letter")
        elif f == "sueldo_bruto_mensual_2":
            bruto_mensual_2_letter = col.get("letter")

    for col in columns:
        f = col.get("field")
        cat = col.get("category")
        formula = col.get("formula")
        letter = col.get("letter")
        
        col["status"] = "direct"
        col["reason"] = "Campo de entrada directa sin fórmula."
        col["recommended_formula"] = None
        
        if f == "vacaciones_restantes":
            expected = f"={tot_letter}6-{tom_letter}6"
            if formula:
                formula_clean = str(formula).upper().replace(" ", "")
                expected_clean = expected.upper().replace(" ", "")
                if formula_clean == expected_clean:
                    col["status"] = "correct"
                    col["reason"] = "Fórmula correcta de vacaciones restantes (restando tomadas de totales)."
                else:
                    col["status"] = "incorrect"
                    col["reason"] = f"Fórmula incorrecta. Se esperaba que restara vacaciones tomadas de las totales: '{expected}'"
                    col["recommended_formula"] = expected
            else:
                col["status"] = "recommended"
                col["reason"] = "Se recomienda calcular automáticamente restando las vacaciones tomadas de las totales."
                col["recommended_formula"] = expected
                
        elif f == "sueldo_bruto_mensual":
            expected = "=SUM(U6:AA6)"
            if formula:
                if str(formula).upper().startswith("=SUM"):
                    col["status"] = "correct"
                    col["reason"] = "Fórmula correcta para sumar conceptos de percepciones."
                else:
                    col["status"] = "incorrect"
                    col["reason"] = f"Se esperaba una sumatoria de percepciones (ej: '{expected}')."
                    col["recommended_formula"] = expected
            else:
                col["status"] = "recommended"
                col["reason"] = "Se recomienda sumar automáticamente todas las percepciones mensuales."
                col["recommended_formula"] = expected
                
        elif f == "sueldo_bruto_quincenal_base":
            expected = f"={bruto_mensual_letter}6/2"
            if formula:
                formula_clean = str(formula).upper().replace(" ", "")
                expected_clean = expected.upper().replace(" ", "")
                if formula_clean == expected_clean:
                    col["status"] = "correct"
                    col["reason"] = "Fórmula correcta de bruto quincenal base (mensual dividido entre 2)."
                else:
                    col["status"] = "incorrect"
                    col["reason"] = f"Fórmula incorrecta. Se esperaba dividir el bruto mensual entre 2: '{expected}'"
                    col["recommended_formula"] = expected
            else:
                col["status"] = "recommended"
                col["reason"] = "Se recomienda dividir el sueldo bruto mensual entre 2."
                col["recommended_formula"] = expected
                
        elif f == "sueldo_bruto_mensual_2":
            ded_letter = "AE"
            for c in columns:
                if c.get("field") == "descuento_adicional":
                    ded_letter = c.get("letter")
            expected = f"={bruto_mensual_letter}6-{ded_letter}6"
            
            if formula:
                formula_clean = str(formula).upper().replace(" ", "")
                expected_clean = expected.upper().replace(" ", "")
                if formula_clean == expected_clean:
                    col["status"] = "correct"
                    col["reason"] = "Fórmula correcta de sueldo neto mensual después de deducciones."
                else:
                    col["status"] = "incorrect"
                    col["reason"] = f"Fórmula incorrecta. Se esperaba restar deducciones del bruto mensual: '{expected}'"
                    col["recommended_formula"] = expected
            else:
                col["status"] = "recommended"
                col["reason"] = "Se recomienda restar deducciones del sueldo bruto mensual."
                col["recommended_formula"] = expected
                
        elif f == "sueldo_bruto_quincenal_2":
            expected = f"={bruto_mensual_2_letter}6/2/15*(15-{faltas_letter}6)"
            if formula:
                formula_clean = str(formula).upper().replace(" ", "")
                expected_clean = expected.upper().replace(" ", "")
                if formula_clean == expected_clean or f"15-{faltas_letter}6" in formula_clean:
                    col["status"] = "correct"
                    col["reason"] = "Fórmula correcta de neto quincenal (descuenta proporcionalmente las faltas)."
                elif "FALTAS" in formula_clean or faltas_letter in formula_clean:
                    col["status"] = "correct"
                    col["reason"] = "Fórmula de neto quincenal con descuento de faltas validada."
                else:
                    col["status"] = "incorrect"
                    col["reason"] = f"Fórmula incorrecta. No descuenta las faltas del periodo. Se recomienda: '{expected}'"
                    col["recommended_formula"] = expected
            else:
                col["status"] = "recommended"
                col["reason"] = "Se recomienda calcular el neto quincenal descontando las faltas del periodo."
                col["recommended_formula"] = expected
                
        elif cat == "calculated":
            if formula:
                col["status"] = "correct"
                col["reason"] = "Columna calculada con fórmula activa."
            else:
                col["status"] = "recommended"
                col["reason"] = "Columna definida como calculada, pero no tiene fórmula activa."
                
    return columns


# ENDPOINTS IN APIHANDLER CLASS
    def schema_validate_endpoint(self, body):
        try:
            excel_path = body.get("path")
            if not excel_path or not os.path.exists(excel_path):
                self.send_json({"error": "No se especificó la ruta del archivo o el archivo no existe."}, 400)
                return

            schema = check_and_heal_schema()
            req_cols = {int(c["index"]): c for c in body.get("columns", [])} if body.get("columns") else {}
            
            # Read spreadsheet headers and formulas
            wb_v = load_workbook_agnostic(excel_path, data_only=True)
            wb_f = load_workbook_agnostic(excel_path, data_only=False)
            sheet_v = wb_v.active
            sheet_f = wb_f.active
            
            headers_row = find_headers_row(sheet_v)
            headers = []
            for col_idx in range(1, sheet_v.max_column + 1):
                val = sheet_v.cell(row=headers_row, column=col_idx).value
                headers.append(val)
                
            wb_v.close()
            
            # Clean trailing nulls in headers
            while headers and (headers[-1] is None or str(headers[-1]).strip() == ""):
                headers.pop()
                
            nombre_idx, sdi_idx = identify_critical_headers(headers)
            if not nombre_idx or not sdi_idx:
                wb_f.close()
                missing = []
                if not nombre_idx: missing.append("Nombre Completo")
                if not sdi_idx: missing.append("SDI / Salario Diario Integrado")
                error_msg = f"El archivo Excel cargado no contiene los campos mínimos obligatorios para el cálculo de la prenómina: {', '.join(missing)}. Por favor, verifica el archivo y vuelve a intentarlo."
                self.send_json({"error": error_msg, "has_minimal_fields": False}, 400)
                return
                
            # Map current columns
            schema_cols = {col.get("index"): col for col in schema.get("columns", [])}
            columns_to_validate = []
            
            for col_idx in range(1, len(headers) + 1):
                header_val = headers[col_idx - 1]
                if header_val is None or str(header_val).strip() == "":
                    continue
                header_str = str(header_val).strip()
                letter = openpyxl.utils.get_column_letter(col_idx)
                
                # Fetch mapped info or guess
                mapped = schema_cols.get(col_idx, {})
                field = mapped.get("field")
                category = mapped.get("category", "metadata")
                label = mapped.get("label", header_str)
                type_val = mapped.get("type", "string")
                editable = mapped.get("editable", True)
                incidence_editable = mapped.get("incidence_editable", False)
                
                if not field:
                    if col_idx == nombre_idx:
                        field = "nombre"
                        category = "metadata"
                        type_val = "string"
                        editable = True
                        incidence_editable = False
                    elif col_idx == sdi_idx:
                        field = "sdi"
                        category = "nominal_imss"
                        type_val = "float"
                        editable = True
                        incidence_editable = False
                    else:
                        import re
                        field = "".join(c for c in header_str.lower() if c.isalnum() or c == " ").strip().replace(" ", "_")
                        if not field:
                            field = f"col_{col_idx}"
                        
                        # Guess category
                        if "vacaciones" in field:
                            if "total" in field or "derecho" in field:
                                field = "vacaciones_totales"
                                category = "metadata"
                                type_val = "float"
                            elif "toma" in field:
                                field = "vacaciones_tomadas"
                                category = "metadata"
                                type_val = "float"
                            elif "resta" in field or "dispon" in field:
                                field = "vacaciones_restantes"
                                category = "calculated"
                                type_val = "float"
                        elif "bruto" in field:
                            if "mensual" in field:
                                field = "sueldo_bruto_mensual"
                                category = "calculated"
                                type_val = "float"
                            elif "quincenal" in field:
                                field = "sueldo_bruto_quincenal_base"
                                category = "calculated"
                                type_val = "float"
                        elif "neto" in field or "quincenal_2" in field:
                            field = "sueldo_bruto_quincenal_2"
                            category = "calculated"
                            type_val = "float"
                        elif "falta" in field:
                            field = "faltas"
                            category = "deduction"
                            type_val = "float"
                            incidence_editable = True
                        elif "retardo" in field:
                            field = "retardos"
                            category = "deduction"
                            type_val = "float"
                            incidence_editable = True
                            
                formula_str = None
                if req_cols and col_idx in req_cols:
                    formula_str = req_cols[col_idx].get("formula")
                else:
                    formula_val = sheet_f.cell(row=6, column=col_idx).value
                    formula_str = str(formula_val) if formula_val and isinstance(formula_val, str) and formula_val.startswith("=") else None
                
                columns_to_validate.append({
                    "index": col_idx,
                    "letter": letter,
                    "header": header_str,
                    "field": field,
                    "type": type_val,
                    "category": category,
                    "label": label,
                    "description": mapped.get("description", f"Columna {header_str}"),
                    "formula": formula_str,
                    "editable": editable,
                    "incidence_editable": incidence_editable
                })
                
            wb_f.close()
            
            # Gemini Call
            gemini_res = None
            api_key = get_ai_api_key(schema.get("ai_provider", "google"))
            if api_key:
                prompt = f"""
                Eres un auditor contable experto en nóminas bajo la LFT (Ley Federal del Trabajo) en México.
                El usuario ha cargado un archivo Excel de nómina. Hemos extraído los encabezados y las fórmulas de la primera fila de datos (renglón 6).
                
                Lista de columnas extraídas:
                {json.dumps(columns_to_validate, ensure_ascii=False, indent=2)}
                
                Parámetros de nómina activos:
                - UMA: {schema.get("uma_cell", "S3")}
                - Días Mes: {schema.get("dias_mes_cell", "N3")}
                
                Por favor, audita el mapeo de campos y valida cada una de las fórmulas según las siguientes reglas:
                1. Clasifica cada columna en un 'status':
                   - "correct" (Verde): Si la fórmula del Excel es correcta y cumple con la LFT y las buenas prácticas.
                   - "incorrect" (Rojo): Si la fórmula tiene un error matemático, conceptual o de sintaxis (ej: divide mal, o calcula el neto quincenal como Sueldo/2 pero el Excel tiene columna de faltas/incidencias y la fórmula la ignora. Debe descontar faltas proporcionalmente).
                   - "recommended" (Azul): Si no tiene fórmula en el Excel (es un valor plano o vacío), pero el Asistente considera que se debería calcular mediante una fórmula (ej: vacaciones_restantes, sueldo_bruto_mensual, sueldo quincenal, neto quincenal, sumatorias).
                   - "direct" (Gris/Plano): Si es un campo que no requiere fórmula y se alimenta directamente (ej: ID, Nombre, Área, Puesto, Salario Diario).
                2. Para cada columna, proporciona:
                   - "description": Una descripción breve y clara de lo que significa este campo.
                   - "reason": Una explicación de por qué tiene ese estado de validación (en español).
                   - "recommended_formula": Una fórmula sugerida si el estado es 'incorrect' o 'recommended' (ej: =AL6-AM6 o =AB6/2). Debe estar en mayúsculas y usar referencias relativas a la fila 6. Si no aplica, ponlo como null.
                
                Responde exclusivamente en formato JSON estructurado con la clave "columns" conteniendo el arreglo de resultados:
                {{
                  "columns": [
                    {{
                      "index": número,
                      "field": "identificador_snake_case",
                      "description": "descripción",
                      "status": "correct|incorrect|recommended|direct",
                      "reason": "explicación",
                      "recommended_formula": "fórmula o null"
                    }}
                  ]
                }}
                """
                try:
                    gemini_res = call_ai_api_simple(prompt, schema, response_json=True)
                except Exception as gem_e:
                    print("Error calling Gemini in validation:", gem_e)
                    
            ai_cols = {c["index"]: c for c in gemini_res["columns"]} if gemini_res and "columns" in gemini_res else {}
            
            final_columns = []
            for col in columns_to_validate:
                idx = col["index"]
                ai_data = ai_cols.get(idx, {})
                
                if ai_data:
                    col["status"] = ai_data.get("status", "direct")
                    col["reason"] = ai_data.get("reason", "Mapeo validado.")
                    col["recommended_formula"] = ai_data.get("recommended_formula")
                    col["description"] = ai_data.get("description", col["description"])
                    if "field" in ai_data:
                        col["field"] = ai_data["field"]
                final_columns.append(col)
                
            if not ai_cols:
                # Fallback to local deterministic validation
                final_columns = validate_schema_formulas_locally(final_columns)
                
            correct_count = sum(1 for c in final_columns if c.get("status") == "correct")
            incorrect_count = sum(1 for c in final_columns if c.get("status") == "incorrect")
            recommended_count = sum(1 for c in final_columns if c.get("status") == "recommended")
            direct_count = sum(1 for c in final_columns if c.get("status") == "direct")
            
            self.send_json({
                "columns": final_columns,
                "summary": {
                    "correct_count": correct_count,
                    "incorrect_count": incorrect_count,
                    "recommended_count": recommended_count,
                    "direct_count": direct_count,
                    "has_minimal_fields": True
                }
            })
        except Exception as e:
            self.send_json({"error": f"Error al validar esquema: {e}", "details": traceback.format_exc()}, 500)

    def schema_confirm_endpoint(self, body):
        try:
            excel_path = body.get("path")
            columns_data = body.get("columns")
            if not excel_path or not os.path.exists(excel_path):
                self.send_json({"error": "Archivo Excel no encontrado."}, 404)
                return
            if not columns_data:
                self.send_json({"error": "No se recibieron las columnas para guardar."}, 400)
                return
                
            schema = load_schema()
            
            clean_columns = []
            for col in columns_data:
                clean_col = {
                    "index": col["index"],
                    "letter": col["letter"],
                    "header": col["header"],
                    "field": col["field"],
                    "type": col.get("type", "string"),
                    "category": col.get("category", "metadata"),
                    "label": col.get("label", col["header"]),
                    "description": col.get("description", ""),
                    "editable": col.get("editable", True),
                    "incidence_editable": col.get("incidence_editable", False)
                }
                if col.get("formula"):
                    clean_col["formula_template"] = col["formula"]
                clean_columns.append(clean_col)
                
            schema["columns"] = clean_columns
            save_schema(schema)
            
            # Open excel and rewrite formulas for row 6 to 20
            wb = load_workbook_agnostic(excel_path, data_only=False)
            ws = wb.active
            headers_row = find_headers_row(ws)
            start_row = headers_row + 1
            
            # Determine end row
            end_row = start_row
            nombre_col = get_field_index(schema, "nombre")
            id_col = get_field_index(schema, "id")
            
            while True:
                n_val = ws.cell(row=end_row, column=nombre_col).value
                c_val = ws.cell(row=end_row, column=id_col).value
                if n_val and any(x in str(n_val).upper() for x in ["TOTAL", "SUMA"]):
                    break
                if n_val is None and c_val is None:
                    break
                end_row += 1
            end_row -= 1
            
            import re
            for r in range(start_row, end_row + 1):
                for col in columns_data:
                    formula = col.get("formula")
                    if formula and str(formula).startswith("="):
                        adapted = re.sub(r'([A-Z]+)6', r'\g<1>' + str(r), formula.upper())
                        ws.cell(row=r, column=col["index"]).value = adapted
                        
            recompile_active_period_incidences(wb, schema)
            save_workbook_agnostic(wb, excel_path)
            wb.close()
            
            self.send_json({"success": True, "message": "Esquema guardado y base de datos recalculada con éxito."})
        except Exception as e:
            self.send_json({"error": f"Error al confirmar esquema: {e}", "details": traceback.format_exc()}, 500)

    def upload_rules_endpoint(self, post_data):
        try:
            boundary = self.headers.get_param('boundary')
            if not boundary:
                self.send_json({"error": "No boundary in Content-Type"}, 400)
                return
            
            parts = parse_multipart(post_data, boundary)
            file_part = parts.get("file")
            if not file_part:
                self.send_json({"error": "No file uploaded in form data"}, 400)
                return
                
            filename = file_part["filename"]
            content = file_part["content"]
            
            if not filename or not content:
                self.send_json({"error": "File is empty or has no name"}, 400)
                return
                
            ext = filename.split('.')[-1].lower()
            if ext == "docx":
                import zipfile
                import xml.etree.ElementTree as ET
                import io
                
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    doc_xml = z.read('word/document.xml')
                    root = ET.fromstring(doc_xml)
                    
                paragraphs = []
                for elem in root.iter():
                    tag_name = elem.tag.split('}')[-1]
                    if tag_name == 'p':
                        p_text = []
                        for child in elem.iter():
                            child_tag = child.tag.split('}')[-1]
                            if child_tag == 't' and child.text:
                                p_text.append(child.text)
                        text_str = "".join(p_text).strip()
                        if text_str:
                            paragraphs.append(text_str)
                extracted_text = "\n\n".join(paragraphs)
                self.send_json({"text": extracted_text})
            elif ext in ["txt", "csv"]:
                try:
                    extracted_text = content.decode("utf-8")
                except UnicodeDecodeError:
                    extracted_text = content.decode("latin-1")
                self.send_json({"text": extracted_text})
            else:
                self.send_json({"error": "Formato de reglas no soportado. Debe ser .docx o .txt"}, 400)
        except Exception as e:
            tb = traceback.format_exc()
            print("Error in upload_rules_endpoint:\n", tb)
            self.send_json({"error": f"Error al leer archivo de reglas: {e}", "details": tb}, 500)

    def get_schema(self):
        schema = check_and_heal_schema()
        self.send_json(schema)

    def download_excel(self):
        try:
            excel_path = get_excel_path()
            if not excel_path or not os.path.exists(excel_path):
                self.send_json({"error": f"Excel file not found at: {excel_path}"}, 404)
                return
            
            filename = os.path.basename(excel_path)
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(os.path.getsize(excel_path)))
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            
            with open(excel_path, "rb") as f:
                self.wfile.write(f.read())
        except Exception as e:
            self.send_json({"error": f"Error downloading file: {str(e)}"}, 500)

    def select_file(self):
        path = select_file_via_dialog()
        self.send_json({"selected_path": path})

    def select_rules_file(self):
        path = select_rules_file_via_dialog()
        if not path:
            self.send_json({"text": None})
            return
            
        try:
            ext = path.split('.')[-1].lower()
            if ext == "docx":
                with open(path, 'rb') as f:
                    data_bytes = f.read()
                import zipfile
                import xml.etree.ElementTree as ET
                import io
                with zipfile.ZipFile(io.BytesIO(data_bytes)) as z:
                    doc_xml = z.read('word/document.xml')
                    root = ET.fromstring(doc_xml)
                    paragraphs = []
                    for elem in root.iter():
                        tag_name = elem.tag.split('}')[-1]
                        if tag_name == 'p':
                            p_text = []
                            for child in elem.iter():
                                child_tag = child.tag.split('}')[-1]
                                if child_tag == 't' and child.text:
                                    p_text.append(child.text)
                            text_str = "".join(p_text).strip()
                            if text_str:
                                paragraphs.append(text_str)
                    extracted_text = "\n\n".join(paragraphs)
            elif ext in ["txt", "md"]:
                with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                    extracted_text = f.read()
            else:
                self.send_json({"error": "Formato de archivo no soportado. Por favor, selecciona un archivo .txt, .md o .docx."}, 400)
                return
                
            self.send_json({"selected_path": path, "text": extracted_text})
        except Exception as e:
            self.send_json({"error": f"Error al leer el archivo de reglas: {e}"}, 500)

    def save_clarify(self, body):
        schema = load_schema()
        field = body.get("field")
        answer = body.get("answer")
        
        # Apply clarification logic (e.g. modify category or custom properties)
        # Find column in schema and update based on user answers
        for col in schema["columns"]:
            if col["field"] == field:
                if "deducciÃ³n" in answer.lower():
                    col["category"] = "deduction"
                    col["incidence_editable"] = True
                    col["type"] = "float"
                elif "percepciÃ³n" in answer.lower():
                    col["category"] = "others"
                    col["editable"] = True
                    col["type"] = "float"
                print(f"Applied clarification for {field}: categorized as {col['category']}")
        
        # Remove from pending list
        if "pending_clarifications" in schema:
            schema["pending_clarifications"] = [q for q in schema["pending_clarifications"] if q.get("field") != field]
        save_schema(schema)
        
        # Rewrite formulas for all employees in Hoja1 to ensure newly categorized deduction/perception is integrated
        excel_path = get_excel_path()
        if os.path.exists(excel_path):
            try:
                wb_re = load_workbook_agnostic(excel_path, data_only=False)
                ws_re = wb_re.active
                nombre_col = get_field_index(schema, "nombre")
                id_col = get_field_index(schema, "id")
                headers_row = find_headers_row(ws_re)
                row = headers_row + 1
                while True:
                    nombre_val = ws_re.cell(row=row, column=nombre_col).value
                    cod_val = ws_re.cell(row=row, column=id_col).value
                    if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                        break
                    if nombre_val is None and cod_val is None:
                        break
                    if nombre_val:
                        inject_formulas_dynamically(ws_re, row, schema)
                    row += 1
                
                # Recompile active period to update the sums of newly categorized columns
                recompile_active_period_incidences(wb_re, schema)
                save_workbook_agnostic(wb_re, excel_path)
                wb_re.close()
                print("Excel formulas and incidences recompiled after clarification of field:", field)
            except Exception as reinj_err:
                print("Error rewriting formulas and compiling on save_clarify:", reinj_err)
                
        self.send_json({"success": True, "schema": schema})

    def get_employees(self):
        try:
            schema = check_and_heal_schema()
            excel_path = get_excel_path()
            copy_template_if_needed(excel_path)
            if not os.path.exists(excel_path):
                self.send_json({"error": f"Database file not found at {excel_path}"}, 500)
                return

            # 1. Recompile incidences in Excel first to ensure calculations are up-to-date for the active period
            try:
                wb_comp = load_workbook_agnostic(excel_path, data_only=False)
                recompile_active_period_incidences(wb_comp, schema)
                save_workbook_agnostic(wb_comp, excel_path)
                wb_comp.close()
            except Exception as comp_err:
                print("Error during Excel recompilation on get_employees:", comp_err)

            # 2. Now load for reading
            wb_v = load_workbook_agnostic(excel_path, data_only=True)
            wb_f = load_workbook_agnostic(excel_path, data_only=False)
            sheet_v = wb_v.active
            sheet_f = wb_f.active

            # Read dynamic configuration cells
            uma_cell_coord = schema.get("uma_cell", "S3")
            vales_pct_cell = schema.get("vales_pct_cell", "P3")
            dias_mes_cell = schema.get("dias_mes_cell", "N3")
            fa_pct_cell = schema.get("fa_pct_cell", "L3")
            aguinaldo_cell = schema.get("aguinaldo_cell", "J3")
            prima_cell = schema.get("prima_cell", "H3")

            def get_cell_float(sheet, cell, default):
                val = sheet[cell].value
                if val is None:
                    return default
                try:
                    return float(str(val).replace(",", "").strip())
                except:
                    return default

            uma = get_cell_float(sheet_v, uma_cell_coord, 117.31)
            vales_pct = get_cell_float(sheet_v, vales_pct_cell, 40.0)
            dias_mes = get_cell_float(sheet_v, dias_mes_cell, 30.4)
            fa_pct = get_cell_float(sheet_v, fa_pct_cell, 11.0)
            aguinaldo = get_cell_float(sheet_v, aguinaldo_cell, 15.0)
            prima = get_cell_float(sheet_v, prima_cell, 25.0)

            config = {
                "uma": uma,
                "valesPct": vales_pct,
                "diasMes": dias_mes,
                "faPct": fa_pct,
                "aguinaldo": aguinaldo,
                "prima": prima
            }

            # 3. Read quincenal incidences into agg dictionary
            period_str = schema.get("period", "16 al 30 Abr 2026")
            start_date, end_date = parse_period_dates(period_str)
            agg = {}
            if "Incidencias" in wb_v.sheetnames:
                ws_inc = wb_v["Incidencias"]
                for r in range(2, ws_inc.max_row + 1):
                    date_val = ws_inc.cell(row=r, column=1).value
                    if not date_val:
                        continue
                    try:
                        row_date = parse_date_robust(date_val)
                        if row_date and start_date <= row_date <= end_date:
                            c_id = clean_employee_id(ws_inc.cell(row=r, column=2).value)
                            if c_id not in agg:
                                agg[c_id] = {
                                    "faltas": 0,
                                    "vacaciones": 0,
                                    "retardos": 0,
                                    "forzar_asistencia": "NO",
                                    "forzar_puntualidad": "NO",
                                    "forzar_vales": "NO",
                                    "ajuste_vales": None,
                                    "ajuste_fondo_ahorro": None
                                }
                            agg[c_id]["faltas"] += int(ws_inc.cell(row=r, column=4).value or 0)
                            agg[c_id]["retardos"] += int(ws_inc.cell(row=r, column=5).value or 0)
                            agg[c_id]["vacaciones"] += int(ws_inc.cell(row=r, column=6).value or 0)
                            
                            max_c = ws_inc.max_column
                            if max_c >= 11 and ws_inc.cell(row=r, column=11).value == "SI":
                                agg[c_id]["forzar_asistencia"] = "SI"
                            if max_c >= 12 and ws_inc.cell(row=r, column=12).value == "SI":
                                agg[c_id]["forzar_puntualidad"] = "SI"
                            if max_c >= 13 and ws_inc.cell(row=r, column=13).value == "SI":
                                agg[c_id]["forzar_vales"] = "SI"
                            
                            if max_c >= 14:
                                aj_val = ws_inc.cell(row=r, column=14).value
                                if aj_val is not None and str(aj_val).strip() != "":
                                    try:
                                        agg[c_id]["ajuste_vales"] = float(aj_val)
                                    except ValueError:
                                        pass
                                        
                            if max_c >= 15:
                                aj_fa = ws_inc.cell(row=r, column=15).value
                                if aj_fa is not None and str(aj_fa).strip() != "":
                                    try:
                                        agg[c_id]["ajuste_fondo_ahorro"] = float(aj_fa)
                                    except ValueError:
                                        pass
                    except:
                        pass

            nombre_col = get_field_index(schema, "nombre")
            id_col = get_field_index(schema, "id")
            
            if nombre_col is None or id_col is None:
                raise ValueError("No se encontraron las columnas críticas 'nombre' (Nombre Completo) o 'id' (Código Checador) en el archivo Excel o en el esquema de mapeo. Verifica las cabeceras de tu archivo.")

            employees = []
            headers_row = find_headers_row(sheet_v)
            row = headers_row + 1
            while True:
                nombre_val = sheet_v.cell(row=row, column=nombre_col).value
                cod_val = sheet_v.cell(row=row, column=id_col).value
                
                if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                    break
                if nombre_val is None and cod_val is None:
                    has_more = False
                    for i in range(1, 4):
                        n = sheet_v.cell(row=row+i, column=nombre_col).value
                        c = sheet_v.cell(row=row+i, column=id_col).value
                        if n or c:
                            has_more = True
                    if not has_more:
                        break
                
                if nombre_val:
                    def val_to_float(cell_val):
                        if cell_val is None: return 0.0
                        v = str(cell_val).replace(",", "").strip()
                        if v in ["-", "", "None"]: return 0.0
                        try: return float(v)
                        except ValueError: return 0.0

                    cod_id = clean_employee_id(cod_val) if cod_val is not None else f"TEMP_{row}"
                    emp_agg = agg.get(cod_id, {
                        "faltas": 0,
                        "vacaciones": 0,
                        "retardos": 0,
                        "forzar_asistencia": "NO",
                        "forzar_puntualidad": "NO",
                        "forzar_vales": "NO",
                        "ajuste_vales": None,
                        "ajuste_fondo_ahorro": None
                    })

                    # Dynamically read all fields defined in schema
                    emp = {
                        "id": cod_id,
                        "_row": row,
                        "_formulas": {},
                        "faltas": emp_agg["faltas"],
                        "vacaciones": emp_agg["vacaciones"],
                        "retardos": emp_agg["retardos"],
                        "forzar_asistencia": emp_agg.get("forzar_asistencia", "NO"),
                        "forzar_puntualidad": emp_agg.get("forzar_puntualidad", "NO"),
                        "forzar_vales": emp_agg.get("forzar_vales", "NO"),
                        "ajuste_vales": emp_agg.get("ajuste_vales"),
                        "ajuste_fondo_ahorro": emp_agg.get("ajuste_fondo_ahorro")
                    }
                    for col in schema["columns"]:
                        f = col["field"]
                        t = col["type"]
                        val = sheet_v.cell(row=row, column=col["index"]).value
                        val_f = sheet_f.cell(row=row, column=col["index"]).value
                        
                        if isinstance(val_f, str) and val_f.startswith("="):
                            emp["_formulas"][f] = val_f
                            
                        if f == "id":
                            continue # Already handled
                        elif t == "float":
                            if f == "vacaciones_restantes":
                                tot_idx = get_field_index(schema, "vacaciones_totales")
                                tom_idx = get_field_index(schema, "vacaciones_tomadas")
                                tot_val = val_to_float(sheet_v.cell(row=row, column=tot_idx).value) if tot_idx else 0.0
                                tom_val = val_to_float(sheet_v.cell(row=row, column=tom_idx).value) if tom_idx else 0.0
                                emp[f] = tot_val - tom_val
                            else:
                                emp[f] = val_to_float(val)
                        elif t == "date":
                            if isinstance(val, datetime):
                                emp[f] = val.strftime("%Y-%m-%d")
                            else:
                                emp[f] = str(val)[:10] if val else ""
                        elif t == "boolean":
                            emp[f] = str(val or "").upper() == "SI"
                        else:
                            emp[f] = str(val or "").strip()
                            
                    employees.append(emp)
                row += 1

            wb_v.close()
            wb_f.close()
            
            self.send_json({
                "period": schema.get("period", "16 al 30 Abr 2026"),
                "uma": uma,
                "config": config,
                "db_path": schema.get("db_path", "Nomina ciega.xlsx"),
                "employees": employees
            })

        except Exception as e:
            tb = traceback.format_exc()
            print("Error get_employees:\n", tb)
            self.send_json({"error": f"Error reading Excel database: {e}", "details": tb}, 500)

    def save_collaborator(self, body):
        try:
            schema = check_and_heal_schema()
            excel_path = get_excel_path()
            if not os.path.exists(excel_path):
                self.send_json({"error": "Database file not found"}, 500)
                return

            cod = body.get("id")
            if not cod:
                self.send_json({"error": "Collaborator ID/Code is required"}, 400)
                return

            wb = load_workbook_agnostic(excel_path, data_only=False)
            ws = wb.active

            nombre_col = get_field_index(schema, "nombre")
            id_col = get_field_index(schema, "id")

            row = 6
            headers_row = find_headers_row(ws)
            row = headers_row + 1
            found_row = None
            totals_row = None
            
            is_temp_id = isinstance(cod, str) and cod.startswith("TEMP_")
            temp_row_resolved = None
            if is_temp_id:
                try:
                    temp_row_resolved = int(cod.split("_")[1])
                except ValueError:
                    pass

            while True:
                nombre_val = ws.cell(row=row, column=nombre_col).value
                cod_val = ws.cell(row=row, column=id_col).value
                
                if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                    totals_row = row
                    break
                if nombre_val is None and cod_val is None:
                    has_more = False
                    for i in range(1, 4):
                        n = ws.cell(row=row+i, column=nombre_col).value
                        c = ws.cell(row=row+i, column=id_col).value
                        if n or c:
                            has_more = True
                    if not has_more:
                        totals_row = row
                        break
                
                if temp_row_resolved == row:
                    found_row = row
                elif not is_temp_id and cod_val is not None and str(cod_val).strip() == str(cod).strip():
                    found_row = row
                row += 1

            if found_row:
                target_row = found_row
                print(f"Updating collaborator {cod} at row {target_row}")
            else:
                target_row = totals_row
                ws.insert_rows(target_row, amount=1)
                print(f"Adding new collaborator {cod} at inserted row {target_row}")

            # Write values dynamically based on schema config
            for col in schema["columns"]:
                f = col["field"]
                if not col.get("editable", True):
                    continue
                t = col["type"]
                val = body.get(f)
                
                cell_ref = ws.cell(row=target_row, column=col["index"])
                if f == "id":
                    cell_ref.value = None if is_temp_id else val
                elif t == "float":
                    v_float = float(val) if val is not None and str(val).strip() != "" else 0.0
                    cell_ref.value = v_float if v_float > 0 else None
                elif t == "date":
                    cell_ref.value = datetime.strptime(val, "%Y-%m-%d") if val else None
                elif t == "boolean":
                    cell_ref.value = "SI" if val else "NO"
                else:
                    cell_ref.value = str(val).strip() if val else None

            # Calculate Factor de Integracion based on current date vs ingreso date
            ingreso_str = body.get("ingreso", "")
            baja_str = body.get("baja", "")
            fi_col = get_field_index(schema, "factor_integracion")
            
            if fi_col:
                if ingreso_str and not baja_str:
                    try:
                        ingreso_dt = datetime.strptime(ingreso_str, "%Y-%m-%d")
                        _, active_date_obj = parse_period_dates(schema.get("period", ""))
                        active_dt = datetime(active_date_obj.year, active_date_obj.month, active_date_obj.day)
                        diff_yrs = (active_dt - ingreso_dt).days / 365.25
                        years = max(1, int(diff_yrs))
                            
                        vac = calculate_vacation_days(years)
                        fi = 1 + (15/365) + ((vac * 0.25) / 365)
                        ws.cell(row=target_row, column=fi_col).value = round(fi, 4)
                    except Exception as ex:
                        print(f"Error calculating Factor Integration: {ex}")
                        ws.cell(row=target_row, column=fi_col).value = 1.0493
                else:
                    ws.cell(row=target_row, column=fi_col).value = 0.0

            # Inject all calculations formulas dynamically
            inject_formulas_dynamically(ws, target_row, schema)

            # Normal neto_quincenal (since no absences on initial creation/edit unless previously set)
            neto_quincenal_col = get_field_index(schema, "neto_quincenal")
            bruto_mensual_neto_letter = get_field_letter(schema, "bruto_mensual_neto")
            if neto_quincenal_col:
                ws.cell(row=target_row, column=neto_quincenal_col).value = f"={bruto_mensual_neto_letter}{target_row}/2"

            # 6. Re-sum totals in Row 21 or new totals row
            new_totals_row = totals_row + 1 if not found_row else totals_row
            
            # Dynamically compile the list of float columns to sum in totals row (excluding rate/daily indicators)
            columns_to_sum = []
            for col in schema.get("columns", []):
                if col.get("type") == "float" and col.get("field") not in ["antiguedad", "factor_integracion", "sdi", "salario_diario"]:
                    columns_to_sum.append(col.get("field"))

            for field in columns_to_sum:
                col_idx = get_field_index(schema, field)
                letter = get_field_letter(schema, field)
                if col_idx and letter:
                    ws.cell(row=new_totals_row, column=col_idx).value = f"=SUM({letter}{headers_row + 1}:{letter}{new_totals_row-1})"
            
            # AC and AG sum
            bruto_quincenal_col = get_field_index(schema, "bruto_quincenal")
            if bruto_quincenal_col:
                ws.cell(row=new_totals_row, column=bruto_quincenal_col).value = f"={bruto_mensual_neto_letter}{new_totals_row}/2"
            if neto_quincenal_col:
                neto_quincenal_letter = get_field_letter(schema, "neto_quincenal")
                ws.cell(row=new_totals_row, column=neto_quincenal_col).value = f"=SUM({neto_quincenal_letter}{headers_row + 1}:{neto_quincenal_letter}{new_totals_row-1})"

            # Save file
            save_workbook_agnostic(wb, excel_path)
            wb.close()
            
            self.send_json({"success": True, "message": f"Collaborator saved at row {target_row}"})
            
        except PermissionError:
            self.send_json({"error": f"El archivo '{os.path.basename(excel_path)}' estÃ¡ abierto en Microsoft Excel o bloqueado por el sistema. Por favor, cierra el archivo local e intÃ©ntalo de nuevo."}, 500)
        except Exception as e:
            self.send_json({"error": f"Error saving collaborator: {e}", "details": traceback.format_exc()}, 500)

    def save_incidences(self, body):
        try:
            schema = check_and_heal_schema()
            excel_path = get_excel_path()
            if not os.path.exists(excel_path):
                self.send_json({"error": "Database file not found"}, 500)
                return

            changes_list = body if isinstance(body, list) else [body]
            if not changes_list:
                self.send_json({"error": "No changes provided"}, 400)
                return

            wb = load_workbook_agnostic(excel_path, data_only=False)
            ws = wb.active

            nombre_col = get_field_index(schema, "nombre")
            id_col = get_field_index(schema, "id")
            headers_row = find_headers_row(ws)

            registered_count = 0
            for change_item in changes_list:
                cod = change_item.get("id") or change_item.get("employee_id")
                if not cod:
                    continue

                row = headers_row + 1
                found_row = None
                is_temp_id = isinstance(cod, str) and cod.startswith("TEMP_")
                temp_row_resolved = None
                if is_temp_id:
                    try:
                        temp_row_resolved = int(cod.split("_")[1])
                    except ValueError:
                        pass

                while True:
                    nombre_val = ws.cell(row=row, column=nombre_col).value
                    cod_val = ws.cell(row=row, column=id_col).value
                    
                    if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                        break
                    if nombre_val is None and cod_val is None:
                        break
                    if temp_row_resolved == row:
                        found_row = row
                        break
                    elif not is_temp_id and cod_val is not None and clean_employee_id(cod_val) == clean_employee_id(cod):
                        found_row = row
                        break
                    row += 1

                if not found_row:
                    if len(changes_list) == 1:
                        self.send_json({"error": f"Collaborator Cód. {cod} not found in database"}, 404)
                        wb.close()
                        return
                    else:
                        print(f"Skipping collaborator Cód. {cod}: not found in database")
                        continue

                nombre_val = ws.cell(row=found_row, column=nombre_col).value

                # Determine appropriate date
                import datetime
                tz_mex = datetime.timezone(datetime.timedelta(hours=-6))
                today_mex = datetime.datetime.now(tz_mex).date()
                today_str = today_mex.strftime("%Y-%m-%d")

                date_val = change_item.get("date") or change_item.get("fecha")
                if not date_val:
                    period_str = schema.get("period", "16 al 30 Abr 2026")
                    start_date, end_date = parse_period_dates(period_str)
                    if not (start_date <= today_mex <= end_date):
                        today_str = start_date.strftime("%Y-%m-%d")
                else:
                    today_str = str(date_val)

                # Save to Incidencias log
                incidence_data = {
                    "date": today_str,
                    "id": cod,
                    "nombre": nombre_val or "",
                    "faltas": int(change_item.get("faltas", 0)),
                    "retardos": int(change_item.get("retardos", 0)),
                    "vacaciones": int(change_item.get("vacaciones", 0)),
                    "descuento_adicional": float(change_item.get("descuento_adicional", 0.0)),
                    "puntualidad": "SI" if change_item.get("puntualidad") in [True, "SI"] else ("NO" if change_item.get("puntualidad") in [False, "NO"] else change_item.get("puntualidad", "SI")),
                    "asistencia": "SI" if change_item.get("asistencia") in [True, "SI"] else ("NO" if change_item.get("asistencia") in [False, "NO"] else change_item.get("asistencia", "SI")),
                    "observaciones": change_item.get("observaciones", ""),
                    "forzar_asistencia": "SI" if change_item.get("forzar_asistencia") in [True, "SI"] else ("NO" if change_item.get("forzar_asistencia") in [False, "NO"] else change_item.get("forzar_asistencia", "NO")),
                    "forzar_puntualidad": "SI" if change_item.get("forzar_puntualidad") in [True, "SI"] else ("NO" if change_item.get("forzar_puntualidad") in [False, "NO"] else change_item.get("forzar_puntualidad", "NO")),
                    "forzar_vales": "SI" if change_item.get("forzar_vales") in [True, "SI"] else ("NO" if change_item.get("forzar_vales") in [False, "NO"] else change_item.get("forzar_vales", "NO")),
                    "ajuste_vales": change_item.get("ajuste_vales"),
                    "ajuste_fondo_ahorro": change_item.get("ajuste_fondo_ahorro")
                }

                # Gather dynamic deductions
                for col in schema.get("columns", []):
                    if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                        field = col.get("field")
                        if field in change_item:
                            try:
                                incidence_data[field] = float(change_item.get(field, 0.0))
                            except (ValueError, TypeError):
                                incidence_data[field] = 0.0

                save_incidence_to_excel(wb, incidence_data)
                registered_count += 1

            if registered_count > 0:
                recompile_active_period_incidences(wb, schema)
                save_workbook_agnostic(wb, excel_path)
            
            wb.close()

            self.send_json({"success": True, "message": f"Incidences registered for {registered_count} entries"})

        except PermissionError:
            self.send_json({"error": f"El archivo '{os.path.basename(excel_path)}' estÃ¡ abierto en Microsoft Excel o bloqueado por el sistema. Por favor, cierra el archivo local e intÃ©ntalo de nuevo."}, 500)
        except Exception as e:
            self.send_json({"error": f"Error saving incidences: {e}", "details": traceback.format_exc()}, 500)

    def save_period(self, body):
        try:
            period = body.get("period", "").strip()
            if not period:
                self.send_json({"error": "El periodo es requerido"}, 400)
                return
            
            schema = load_schema()
            schema["period"] = period
            save_schema(schema)
            
            # Recompile Hoja1 calculations based on the new active period
            excel_path = get_excel_path()
            if os.path.exists(excel_path):
                try:
                    wb = load_workbook_agnostic(excel_path, data_only=False)
                    recompile_active_period_incidences(wb, schema)
                    save_workbook_agnostic(wb, excel_path)
                    wb.close()
                except Exception as ex_err:
                    print("Error compiling active period incidences on save_period:", ex_err)
            
            self.send_json({"success": True, "message": f"Periodo actualizado a: {period}"})
            
        except Exception as e:
            self.send_json({"error": f"Error saving period: {e}", "details": traceback.format_exc()}, 500)

    def save_config(self, body):
        try:
            uma = float(body.get("uma", 117.31))
            vales_pct = float(body.get("vales_pct", 40.0))
            dias_mes = float(body.get("dias_mes", 30.4))
            fa_pct = float(body.get("fa_pct", 11.0))
            aguinaldo = float(body.get("aguinaldo", 15.0))
            prima = float(body.get("prima", 25.0))
            ai_provider = body.get("ai_provider", "google").strip().lower()
            ai_model = upgrade_model_name(body.get("ai_model", "gemini-2.5-flash"))
            db_path = body.get("db_path", "")
            payroll_rules = body.get("payroll_rules", "")

            schema = load_schema()
            schema["ai_provider"] = ai_provider
            schema["ai_model"] = ai_model
            # Remove legacy gemini_api_key from schema if present
            schema.pop("gemini_api_key", None)
            schema["uma_cell"] = "S3"
            schema["vales_pct_cell"] = "P3"
            schema["dias_mes_cell"] = "N3"
            schema["fa_pct_cell"] = "L3"
            schema["aguinaldo_cell"] = "J3"
            schema["prima_cell"] = "H3"
            schema["payroll_rules"] = payroll_rules
            
            if db_path is not None:
                db_path_lower = db_path.strip().lower()
                if db_path_lower.endswith(".pages") or db_path_lower.endswith(".numbers"):
                    self.send_json({"error": "Formato de archivo no soportado. Por favor usa Excel (.xlsx) o CSV (.csv)."}, 400)
                    return
                schema["db_path"] = db_path.strip()
                
            save_schema(schema)

            excel_path = get_excel_path()
            copy_template_if_needed(excel_path)

            wb = load_workbook_agnostic(excel_path, data_only=False)
            ws = wb.active

            # Write configurations and their labels in Row 3 of Excel
            ws["G3"].value = "PRIMA %:"
            ws["H3"].value = prima
            ws["H3"].number_format = 'General'
            ws["I3"].value = "AGUINALDO DYS:"
            ws["J3"].value = aguinaldo
            ws["J3"].number_format = 'General'
            ws["K3"].value = "FA %:"
            ws["L3"].value = fa_pct
            ws["L3"].number_format = 'General'
            ws["M3"].value = "DIAS MES:"
            ws["N3"].value = dias_mes
            ws["N3"].number_format = 'General'
            ws["O3"].value = "VALES %:"
            ws["P3"].value = vales_pct
            ws["P3"].number_format = 'General'
            ws["R3"].value = "UMA 2026:"
            ws["S3"].value = uma
            ws["S3"].number_format = 'General'

            save_workbook_agnostic(wb, excel_path)
            wb.close()

            self.send_json({"success": True, "message": "Global configuration saved in Excel."})

        except PermissionError:
            self.send_json({"error": f"El archivo '{os.path.basename(excel_path)}' estÃ¡ abierto en Microsoft Excel o bloqueado por el sistema. Por favor, cierra el archivo local e intÃ©ntalo de nuevo."}, 500)
        except Exception as e:
            self.send_json({"error": f"Error saving global configuration: {e}"}, 500)

    def explain_payroll(self, body):
        try:
            schema = check_and_heal_schema()
            excel_path = get_excel_path()
            if not os.path.exists(excel_path):
                self.send_json({"error": "Database file not found"}, 500)
                return

            cod = body.get("employee_id")
            is_global = not cod or str(cod).strip().upper() in ["GLOBAL", "NONE", "NULL", ""]

            chat_history = body.get("chat_history", [])
            new_message = body.get("new_message", "")

            wb_v = load_workbook_agnostic(excel_path, data_only=True)
            wb_f = load_workbook_agnostic(excel_path, data_only=False)
            sheet_v = wb_v.active
            sheet_f = wb_f.active

            uma_cell_coord = schema.get("uma_cell", "S3")
            vales_pct_cell = schema.get("vales_pct_cell", "P3")
            dias_mes_cell = schema.get("dias_mes_cell", "N3")
            fa_pct_cell = schema.get("fa_pct_cell", "L3")
            aguinaldo_cell = schema.get("aguinaldo_cell", "J3")
            prima_cell = schema.get("prima_cell", "H3")

            def get_cell_float(sheet, cell, default):
                val = sheet[cell].value
                if val is None:
                    return default
                try:
                    return float(str(val).replace(",", "").strip())
                except:
                    return default

            uma = get_cell_float(sheet_v, uma_cell_coord, 117.31)
            vales_pct = get_cell_float(sheet_v, vales_pct_cell, 40.0)
            dias_mes = get_cell_float(sheet_v, dias_mes_cell, 30.4)
            fa_pct = get_cell_float(sheet_v, fa_pct_cell, 11.0)
            aguinaldo = get_cell_float(sheet_v, aguinaldo_cell, 15.0)
            prima = get_cell_float(sheet_v, prima_cell, 25.0)

            nombre_col = get_field_index(schema, "nombre")
            id_col = get_field_index(schema, "id")
            headers_row = find_headers_row(sheet_v)

            companies = load_companies()
            companies_map = {c["nombre"].strip().upper(): c for c in companies}

            period_str = schema.get("period", "16 al 30 Abr 2026")
            start_date, end_date = parse_period_dates(period_str)
            incidences_map = {}

            if "Incidencias" in wb_v.sheetnames:
                ws_inc = wb_v["Incidencias"]
                for r in range(2, ws_inc.max_row + 1):
                    date_val = ws_inc.cell(row=r, column=1).value
                    if date_val:
                        try:
                            import datetime
                            if isinstance(date_val, datetime.date) or isinstance(date_val, datetime.datetime):
                                row_date = date_val.date() if isinstance(date_val, datetime.datetime) else date_val
                            else:
                                row_date = datetime.datetime.strptime(str(date_val)[:10], "%Y-%m-%d").date()
                            
                            if start_date <= row_date <= end_date:
                                c_id = clean_employee_id(ws_inc.cell(row=r, column=2).value)
                                if c_id:
                                    if c_id not in incidences_map:
                                        incidences_map[c_id] = 0
                                    incidences_map[c_id] += int(ws_inc.cell(row=r, column=4).value or 0)
                        except:
                            pass

            def val_to_float(cell_val):
                if cell_val is None: return 0.0
                v = str(cell_val).replace(",", "").strip()
                if v in ["-", "", "None"]: return 0.0
                try: return float(v)
                except ValueError: return 0.0

            all_employees_data = []
            row = headers_row + 1
            while True:
                nombre_val = sheet_v.cell(row=row, column=nombre_col).value
                cod_val = sheet_v.cell(row=row, column=id_col).value
                
                if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                    break
                if nombre_val is None and cod_val is None:
                    break
                
                emp_id_str = clean_employee_id(cod_val) if cod_val is not None else f"TEMP_{row}"
                
                if not is_global and emp_id_str != clean_employee_id(cod):
                    row += 1
                    continue
                
                emp_data = {}
                for col in schema["columns"]:
                    f = col["field"]
                    t = col["type"]
                    val = sheet_v.cell(row=row, column=col["index"]).value
                    if t == "float":
                        if f == "vacaciones_restantes":
                            tot_idx = get_field_index(schema, "vacaciones_totales")
                            tom_idx = get_field_index(schema, "vacaciones_tomadas")
                            tot_val = val_to_float(sheet_v.cell(row=row, column=tot_idx).value) if tot_idx else 0.0
                            tom_val = val_to_float(sheet_v.cell(row=row, column=tom_idx).value) if tom_idx else 0.0
                            emp_data[f] = tot_val - tom_val
                        else:
                            emp_data[f] = val_to_float(val)
                    elif t == "boolean":
                        emp_data[f] = str(val or "").upper() == "SI"
                    else:
                        emp_data[f] = str(val or "").strip()

                emp_faltas = 0
                if emp_id_str in incidences_map:
                    emp_faltas = incidences_map[emp_id_str]
                else:
                    neto_quincenal_col = get_field_index(schema, "neto_quincenal")
                    if neto_quincenal_col:
                        formula_ag = sheet_f.cell(row=row, column=neto_quincenal_col).value
                        if isinstance(formula_ag, str) and "/15*" in formula_ag:
                            try:
                                parts = formula_ag.split("*")
                                days_worked = int(parts[-1])
                                emp_faltas = 15 - days_worked
                            except:
                                pass
                emp_data["faltas"] = emp_faltas
                all_employees_data.append((row, emp_data))
                
                if not is_global:
                    break
                row += 1

            wb_v.close()
            wb_f.close()

            if not all_employees_data:
                if not is_global:
                    self.send_json({"error": f"Colaborador con Cód. {cod} no encontrado"}, 404)
                else:
                    self.send_json({"error": "No se encontraron colaboradores en la base de datos"}, 404)
                return

            custom_rules = schema.get("payroll_rules", "").strip()
            rules_source = "custom" if custom_rules else "official"
            
            if custom_rules:
                rules_to_use = f"Reglas privadas/particulares de la empresa:\n{custom_rules}"
            else:
                rules_to_use = (
                    "No se configuraron reglas privadas de la empresa. Se aplican las reglas contables oficiales estándar de la LFT (Ley Federal del Trabajo) en México:\n"
                    "- Factor de Integración = 1 + (Aguinaldo / 365) + (Vacaciones * Prima_Vacacional / 365)\n"
                    "- Salario Diario Integrado (SDI) = Salario Diario * Factor de Integración\n"
                    "- Sueldo Nominal Mensual = Salario Diario * Días del Mes (Row 3)\n"
                    "- Premios de Asistencia y Puntualidad: 10% del SDI * Días del Mes cada uno\n"
                    "- Vales de Despensa = UMA * Vales% * Días del Mes\n"
                    "- Fondo de Ahorro = Sueldo Nominal * FA% (si aplica)\n"
                    "- Sueldo Bruto Mensual = Total Percepciones + Otros Ingresos\n"
                    "- Sueldo Bruto Quincenal normal = Bruto Mensual / 2\n"
                    "- Descuento de Faltas proporcional = (Bruto Quincenal / 15) * Faltas\n"
                    "- Sueldo Neto Quincenal = (Sueldo Bruto Mensual - Descuento Adicional - Deuda Carro) / 2 / 15 * (15 - Faltas)"
                )

            if not is_global:
                found_row, emp_data = all_employees_data[0]
                nombre = emp_data.get("nombre", "Sin Nombre")
                salario_diario = emp_data.get("salario_diario", 0.0)
                fi = emp_data.get("factor_integracion", 1.0493)
                sdi = emp_data.get("sdi", 0.0)
                sueldo_nominal = emp_data.get("sueldo_nominal", 0.0)
                puntualidad = emp_data.get("puntualidad", 0.0)
                asistencia = emp_data.get("asistencia", 0.0)
                vales_despensa = emp_data.get("vales_despensa", 0.0)
                fondo_ahorro = emp_data.get("fondo_ahorro", 0.0)
                fondo_ahorro_activo = "SI" if emp_data.get("fondo_ahorro_activo", False) else "NO"
                percepcion_sueldos = emp_data.get("percepcion_sueldos", 0.0)
                
                asimilados = emp_data.get("asimilados", 0.0)
                gasolina = emp_data.get("gasolina", 0.0)
                socio = emp_data.get("socio", 0.0)
                efectivo = emp_data.get("efectivo", 0.0)
                facturado = emp_data.get("facturado", 0.0)
                deuda_carro = emp_data.get("deuda_carro", 0.0)
                total_otros = asimilados + gasolina + socio + efectivo + facturado + deuda_carro
                bruto_mensual = percepcion_sueldos + total_otros
                bruto_quincenal = bruto_mensual / 2
                
                faltas = emp_data.get("faltas", 0)
                descuento_faltas = (bruto_quincenal / 15.0) * faltas if faltas > 0 else 0.0
                descuento_adicional = emp_data.get("descuento_adicional", 0.0)
                neto_quincenal = max(0.0, (bruto_mensual - descuento_adicional) / 2 / 15 * (15 - faltas))

                ingreso_str = emp_data.get("ingreso", "")
                years_of_labores = 0.0
                vac = 12
                if ingreso_str:
                    try:
                        ingreso_dt = datetime.datetime.strptime(ingreso_str, "%Y-%m-%d")
                        _, active_date_obj = parse_period_dates(schema.get("period", ""))
                        active_dt = datetime.datetime(active_date_obj.year, active_date_obj.month, active_date_obj.day)
                        years_of_labores = (active_dt - ingreso_dt).days / 365.25
                        y = max(1, int(years_of_labores))
                        vac = calculate_vacation_days(y)
                    except:
                        pass
                
                emp_company_name = emp_data.get("empresa", "").strip().upper()
                emp_company = companies_map.get(emp_company_name, {
                    "nombre": emp_data.get("empresa", ""),
                    "razon_social": emp_data.get("empresa", ""),
                    "regimen": "Régimen General de Ley Personas Morales (Por defecto)",
                    "prima_riesgo": 0.5432
                })

                fa_status = f"Sí, activo (`={fa_pct:.0f}%`): `=Sueldo_Nominal * {fa_pct / 100:.2f}` que equivale a **${fondo_ahorro:,.2f}**" if (fondo_ahorro_activo == "SI" and fondo_ahorro > 0) else "No activo"
                
                local_desglose = f"""### 📝 Explicación del Cálculo de Nómina (Offline)

*Nota: No hay una clave de API de Gemini válida configurada en la base de datos, por lo que se muestra el desglose matemático contable estándar.*

**Colaborador:** {nombre} (Código: {cod})  
**Empresa:** {emp_company.get('nombre')} ({emp_company.get('razon_social')})
**Régimen Fiscal:** {emp_company.get('regimen')} | **Prima de Riesgo:** {emp_company.get('prima_riesgo')}%
**Fecha de Ingreso:** {ingreso_str}  
**Antigüedad:** {years_of_labores:.2f} años ({vac} días de vacaciones correspondientes según la LFT)  

---

#### 1. Esquema Nominal IMSS (Base Fiscal)
* **Factor de Integración (FI):**  
  Fórmula Excel: `=1 + (Días_Aguinaldo / 365) + (Días_Vacaciones * Prima_Vacacional / 365)`  
  Cálculo: `=1 + ({aguinaldo:.0f} / 365) + ({vac} * {prima / 100:.2f} / 365)`  
  Resultado: `{fi:.4f}`
* **Salario Diario Integrado (SDI):**  
  Fórmula Excel: `=Salario_Diario * Factor_Integracion`  
  Cálculo: `=${salario_diario:,.2f} * {fi:.4f}`  
  Resultado: **${sdi:,.2f}** (Base de cotización ante el IMSS)
* **Sueldo Nominal Mensual:**  
  Fórmula Excel: `=Salario_Diario * Días_del_Mes`  
  Cálculo: `=${salario_diario:,.2f} * {dias_mes:.1f}`  
  Resultado: **${sueldo_nominal:,.2f}**
* **Premios de Asistencia y Puntualidad (10% del SDI mensual cada uno):**  
  * **Puntualidad:** **${puntualidad:,.2f}** (Fórmula Excel: `=SDI * 10% * Días_del_Mes` ➡️ `=${sdi:,.2f} * 0.10 * {dias_mes:.1f}`)  
  * **Asistencia:** **${asistencia:,.2f}** (Fórmula Excel: `=SDI * 10% * Días_del_Mes` ➡️ `=${sdi:,.2f} * 0.10 * {dias_mes:.1f}`)  
* **Vales de Despensa:**  
  Fórmula Excel: `=UMA * Porcentaje_Vales * Días_del_Mes`  
  Cálculo: `=${uma:,.2f} * {vales_pct / 100:.2f} * {dias_mes:.1f}`  
  Resultado: **${vales_despensa:,.2f}**
* **Fondo de Ahorro:** {fa_status}
* **Total Percepciones Mensuales:** **${percepcion_sueldos:,.2f}**

---

#### 2. Otros Conceptos (Esquema Mixto / Adicionales)
* Honorarios Asimilados: `${asimilados:,.2f}` (Mensual)
* Gasolina/Combustible: `${gasolina:,.2f}` (Mensual)
* Pago Socio: `${socio:,.2f}` (Mensual)
* Pago en Efectivo: `${efectivo:,.2f}` (Mensual)
* Pago Facturado: `${facturado:,.2f}` (Mensual)
* Abono Carro: `${deuda_carro:,.2f}` (Mensual)
* **Total de Otros Conceptos:** **${total_otros:,.2f}** (Mensual)

---

#### 3. Cálculo de Prenómina Quincenal (Pago Actual)
* **Sueldo Bruto Mensual (Base Total):** **${bruto_mensual:,.2f}** (Sueldo Nominal + Otros Conceptos)
* **Sueldo Bruto Quincenal:** **${bruto_quincenal:,.2f}** (Fórmula Excel: `=Sueldo_Bruto_Mensual / 2`)
* **Ajustes, Faltas y Descuentos en la Quincena:**
  * **Descuento por Faltas ({faltas} días):** Deducción de **${descuento_faltas:,.2f}**
  * **Descuento Adicional:** **${descuento_adicional:,.2f}**
* **Sueldo Neto Quincenal Final a Pagar:**  
  Fórmula Excel: `=(Sueldo_Bruto_Mensual - Descuento_Adicional) / 2 / 15 * (15 - Faltas)`  
  Cálculo: `=(${bruto_mensual:,.2f} - ${descuento_adicional:,.2f}) / 2 / 15 * (15 - {faltas})`  
  Resultado: **${neto_quincenal:,.2f}**
"""
                collab_details = f"""Por favor, genera la explicación inicial y detallada del cálculo de prenómina para este colaborador:
DATOS DEL COLABORADOR Y SU EMPRESA:
- Nombre: {nombre}
- Código: {cod}
- Empresa: {emp_company.get('nombre')} (Razón Social: {emp_company.get('razon_social')})
- Régimen Fiscal de la Empresa: {emp_company.get('regimen')}
- Prima de Riesgo de la Empresa: {emp_company.get('prima_riesgo')}%
- Salario Diario: ${salario_diario:,.2f}
- Fecha de Ingreso: {ingreso_str}
- Antigüedad: {years_of_labores:.2f} años ({vac} días de vacaciones según LFT)
- Cuenta con Fondo de Ahorro: {fondo_ahorro_activo}
- Faltas registradas en la quincena: {faltas}

VALORES REGISTRADOS EN EXCEL:
- Salario Diario: ${salario_diario:,.2f}
- Factor de Integración: {fi:.4f}
- Salario Diario Integrado (SDI): ${sdi:,.2f}
- Sueldo Nominal Mensual: ${sueldo_nominal:,.2f}
- Premio de Puntualidad Mensual: ${puntualidad:,.2f}
- Premio de Asistencia Mensual: ${asistencia:,.2f}
- Vales de Despensa Mensual: ${vales_despensa:,.2f}
- Fondo de Ahorro Mensual: ${fondo_ahorro:,.2f}
- Total Percepciones Mensual: ${percepcion_sueldos:,.2f}
- Asimilados Mensual: ${asimilados:,.2f}
- Gasolina Mensual: ${gasolina:,.2f}
- Pago Socio Mensual: ${socio:,.2f}
- Efectivo Mensual: ${efectivo:,.2f}
- Facturado Mensual: ${facturado:,.2f}
- Abono Carro Mensual: ${deuda_carro:,.2f}
- Total Otros Ingresos Mensual: ${total_otros:,.2f}
- Sueldo Bruto Mensual: ${bruto_mensual:,.2f}
- Sueldo Bruto Quincenal base: ${bruto_quincenal:,.2f}
- Descuento por Faltas en Quincena: ${descuento_faltas:,.2f}
- Descuento Adicional en Quincena: ${descuento_adicional:,.2f}
- Sueldo Neto Quincenal Final: ${neto_quincenal:,.2f}
"""
            # Get active dynamic deductions to include in prompt
            ded_cols = [col for col in schema.get("columns", []) if col.get("category") == "deduction" and col.get("incidence_editable")]
            dynamic_ded_fields_str = ""
            for col in ded_cols:
                if col.get("field") != "descuento_adicional":
                    dynamic_ded_fields_str += f'\n    "{col.get("field")}": 0.0,     // (opcional) {col.get("label") or col.get("header")}'

            # Define system prompt
            system_prompt = f"""Eres el Asistente AI de Prenómina de RHM. Tu función es ayudar al administrador a entender, validar y realizar modificaciones en el cálculo de nómina de los colaboradores de forma concisa y directa.

Normativa y Reglas de Nómina aplicables:
{rules_to_use}

Deducciones de Ley Estimadas (para tu explicación contable):
- Dependiendo del régimen fiscal de la empresa a la que pertenece el colaborador, estima los impuestos de la siguiente manera:
  * Si el régimen es "Sueldos y Salarios / Asimilados": NO se descuenta/retiene IMSS Obrero. El ISR quincenal se estima de forma progresiva: si la quincena es menor o igual a $3,700 es 6%; entre $3,701 y $7,000 es 10%; entre $7,001 y $12,000 es 16%; y mayor a $12,000 es 20%.
  * Si el régimen es "RESICO Personas Físicas": Se retiene IMSS Obrero normal (2.375% sobre el SDI multiplicado por los días laborados reales: `SDI * 2.375% * (15 - faltas)`). El ISR es una tasa fija del 1.25% sobre el sueldo bruto quincenal.
  * Si el régimen es "Régimen General de Ley Personas Morales": Se retiene IMSS Obrero normal (2.375% sobre el SDI multiplicado por los días laborados reales: `SDI * 2.375% * (15 - faltas)`). El ISR quincenal se estima de forma progresiva (6%, 10%, 16%, 20%). La Prima de Riesgo de la empresa afecta el costo de seguridad social patronal total.
- Siempre incluye un desglose estimado muy breve de estas deducciones fiscales de ley (ISR e IMSS) y el Neto quincenal estimado resultante en tu explicación. Si la incidencia es de una fecha pasada fuera del periodo de pago activo, aclara brevemente que no afectará el neto de esta quincena. Evita explicaciones extensas o redundantes de discrepancias.

Instrucciones Especiales de Lógica Contable:
- Condonaciones o Justificaciones: Si el usuario te indica que una falta está justificada o condonada (ej. "trajo incapacidad", "perdónale la falta", "págale completo"), debes generar un JSON de cambios estableciendo `"forzar_asistencia": "SI"`, `"forzar_vales": "SI"` y `"forzar_puntualidad": "SI"` (según corresponda), y escribir el motivo en `"observaciones"` (ej. "Incapacidad médica justificante / Faltas condonadas").
- Planes de Amortización de Préstamos: Si el usuario te indica registrar un préstamo de $M para pagarse en N quincenas, divide el monto (M / N) y genera un JSON con el campo de deducción correspondiente (ej. `"descuento_adicional"`) establecido a ese monto quincenal (redondeado a centavos). Escribe en `"observaciones"` el desglose descriptivo de la amortización, ej. "Amortización de préstamo quincena 1 de N (monto quincenal: $Q, total: $M)".
- Registro de Vacaciones Pasadas (Históricas): Si el usuario indica que un colaborador ya tomó N días de vacaciones antes del uso de este sistema (ej. "ya tomó 3 días de vacaciones en enero"), debes generar un JSON estableciendo `"vacaciones": N`, `"date": "YYYY-MM-DD"` (escribe la fecha exacta mencionada, o una fecha representativa del pasado dentro de su ciclo de aniversario actual, por ejemplo en enero u otro mes según corresponda), y `"observaciones": "Registro histórico: Vacaciones tomadas previas al uso del sistema"`. El sistema acumulará estas vacaciones para descontarlas del derecho anual disponible (vacaciones restantes), pero NUNCA debes restar días de nómina ni aplicar deducciones/descuentos económicos en la quincena actual por estas incidencias pasadas.
- Incidencias de Fechas Pasadas (Generales): Cualquier incidencia (vacaciones, faltas, retardos) cuya fecha (`"date"`) esté fuera del período activo de la nómina actual se registra puramente como historial en la pestaña de incidencias y para actualizar saldos de vacaciones. NUNCA debes aplicar descuentos de sueldo, deducciones económicas ni reducciones de días laborados de la quincena actual por incidencias de periodos pasados. En tu explicación, confirma que se registra históricamente en el sistema sin alterar la nómina actual.

Instrucciones Generales:
1. Explica el desglose o responde la duda de forma sumamente concisa, clara, directa y profesional usando Markdown.
2. Si el usuario te da instrucciones de registrar incidencias, modificar datos o aplicar ajustes (ej. "Tiene 2 faltas", "Descuéntale tanto a un empleado", "Un empleado tuvo dos faltas", etc.), limita tu respuesta a una confirmación muy breve y directa, y coloca de inmediato al final de tu respuesta el bloque de código JSON con los cambios en `"apply_changes"`.
3. Si el usuario te pide filtrar o buscar empleados con ciertas características (ej. "Dime quiénes no tienen fondo de ahorro y darmelos y filtrar la tabla", "Muéstrame a los de BYRMAX", "Quiénes tienen faltas?", etc.), responde enumerando a dichos empleados y coloca al final de tu respuesta el bloque de código JSON con la lista de IDs bajo `"filter_employee_ids"`.
4. Si el usuario solo está conversando o haciendo preguntas sin solicitar un cambio o filtro, responde de forma concisa y NO incluyas el bloque JSON.

Formatos del bloque JSON (debe ser un bloque de código Markdown con ```json ... ```):

A) Para filtrar la tabla (por ejemplo, al preguntar "quiénes no tienen fondo de ahorro"):
{{
  "filter_employee_ids": ["190", "171", "108"]
}}

B) Para aplicar cambios a un único empleado específico (por ejemplo, "descuéntale 500 a Juan" o "el empleado 190 tuvo 2 faltas"):
{{
  "apply_changes": {{
    "id": "190",                       // Cód. del empleado (Obligatorio en consultas globales/búsquedas por nombre)
    "date": "2026-06-02",             // (opcional) fecha en formato YYYY-MM-DD
    "faltas": 2,                      // (opcional) número total de faltas en la quincena
    "retardos": 1,                    // (opcional) número total de retardos
    "vacaciones": 3,                  // (opcional) número total de vacaciones tomadas
    "descuento_adicional": 500.0,     // (opcional) monto del descuento adicional en pesos{dynamic_ded_fields_str}
    "observaciones": "Texto...",       // (opcional) observaciones sobre la incidencia o cambio
    "puntualidad": false,             // (opcional) true/false o "SI"/"NO" para el bono de puntualidad
    "asistencia": false,              // (opcional) true/false o "SI"/"NO" para el bono de asistencia
    "forzar_asistencia": "SI",        // (opcional) "SI"/"NO"
    "forzar_puntualidad": "SI",       // (opcional) "SI"/"NO"
    "forzar_vales": "SI",             // (opcional) "SI"/"NO"
    "ajuste_vales": 1200.0,           // (opcional) monto fijo para sobreescribir vales de despensa
    "ajuste_fondo_ahorro": 800.0      // (opcional) monto fijo para sobreescribir fondo de ahorro
  }}
}}

C) Para aplicar cambios a múltiples empleados a la vez (por ejemplo, "Empleado 190 tiene 2 faltas y Empleado 171 tiene 1"):
{{
  "apply_changes": [
    {{
      "id": "190",
      "faltas": 2
    }},
    {{
      "id": "171",
      "faltas": 1
    }}
  ]
}}

Nota: El descuento por faltas se calculará automáticamente con base en las faltas registradas. El descuento adicional es acumulativo para otros conceptos puntuales.
"""

            # Build conversation contents
            contents = []
            for msg in chat_history:
                role = "user" if msg.get("role") in ["user", "role"] else "model"
                text_content = msg.get("content") or msg.get("text") or ""
                contents.append({
                    "role": role,
                    "parts": [{"text": text_content}]
                })
            
            # Add user query
            if not chat_history:
                # If first message, prepend employee info
                if is_global:
                    # Global query
                    emp_list_desc = ""
                    for r_num, emp_d in all_employees_data:
                        fa_status = "SI" if emp_d.get("fondo_ahorro_activo") else "NO"
                        emp_list_desc += (
                            f"- {emp_d.get('nombre')} (Cód: {emp_d.get('id') or ''}): "
                            f"Empresa: {emp_d.get('empresa')}, "
                            f"Régimen: {companies_map.get(emp_d.get('empresa', '').strip().upper(), {}).get('regimen', 'Régimen General de Ley Personas Morales')}, "
                            f"Prima de Riesgo: {companies_map.get(emp_d.get('empresa', '').strip().upper(), {}).get('prima_riesgo', 0.5432)}%, "
                            f"Salario Diario: ${emp_d.get('salario_diario', 0.0):,.2f}, "
                            f"Faltas: {emp_d.get('faltas', 0)}, "
                            f"Retardos: {emp_d.get('retardos', 0)}, "
                            f"Cuenta con Fondo de Ahorro: {fa_status}, "
                            f"Sueldo Neto Quincenal: ${emp_d.get('neto_quincenal', 0.0):,.2f}, "
                            f"Descuento Adicional: ${emp_d.get('descuento_adicional', 0.0):,.2f}, "
                            f"Vacaciones Restantes: {emp_d.get('vacaciones_restantes', 0.0):.1f}, "
                            f"Antigüedad: {emp_d.get('antiguedad', 0.0):.2f} años, "
                            f"Observaciones: {emp_d.get('observaciones', '')}\n"
                        )
                    
                    contents.append({
                        "role": "user",
                        "parts": [{"text": f"Contexto de colaboradores registrados:\n{emp_list_desc}\n\nPregunta/Instrucción: {new_message}"}]
                    })
                else:
                    contents.append({
                        "role": "user",
                        "parts": [{"text": f"{collab_details}\n\nPregunta/Instrucción: {new_message}"}]
                    })
            else:
                contents.append({
                    "role": "user",
                    "parts": [{"text": new_message}]
                })

            # Call AI Chat API (supports Google and OpenRouter)
            try:
                text = call_ai_api_chat(system_prompt, contents, schema)
                if not text:
                    raise Exception("AI API returned empty response")
                    
                # Parse apply_changes from response text
                applied_changes = False
                proposed_changes = None
                filter_employee_ids = None
                try:
                    import re
                    # Robustly try to find a JSON block even if missing ```json
                    json_str = None
                    m = re.search(r"```(?:json)?\s*(.*?)\s*```", text, re.DOTALL)
                    if m:
                        json_str = m.group(1).strip()
                    else:
                        m_direct = re.search(r"(\{.*?(?:\"apply_changes\"|\"filter_employee_ids\").*?\})", text, re.DOTALL)
                        if m_direct:
                            json_str = m_direct.group(1).strip()
                    
                    if json_str:
                        changes_data = json.loads(json_str)
                        if "filter_employee_ids" in changes_data:
                            filter_employee_ids = changes_data["filter_employee_ids"]
                            if not isinstance(filter_employee_ids, list):
                                filter_employee_ids = [str(filter_employee_ids)]
                            else:
                                filter_employee_ids = [str(fid) for fid in filter_employee_ids]
                        
                        if "apply_changes" in changes_data:
                            changes = changes_data["apply_changes"]
                            changes_list_to_process = changes if isinstance(changes, list) else [changes]
                            proposed_changes = []
                            
                            wb = load_workbook_agnostic(excel_path, data_only=False)
                            ws = wb.active
                            
                            nombre_col = get_field_index(schema, "nombre")
                            id_col = get_field_index(schema, "id")
                            headers_row = find_headers_row(ws)
                            
                            import datetime
                            tz_mex = datetime.timezone(datetime.timedelta(hours=-6))
                            today_mex = datetime.datetime.now(tz_mex).date()
                            period_str = schema.get("period", "16 al 30 Abr 2026")
                            start_date, end_date = parse_period_dates(period_str)
                            
                            ws_inc = wb["Incidencias"] if "Incidencias" in wb.sheetnames else None
                            if ws_inc:
                                heal_incidences_sheet_if_needed(wb, schema)
                                
                            for change_item in changes_list_to_process:
                                # Resolve target employee code
                                target_cod = change_item.get("id") or change_item.get("employee_id") or (None if is_global else cod)
                                if not target_cod:
                                    continue
                                
                                # Find employee row
                                row = headers_row + 1
                                found_row = None
                                is_temp_id = isinstance(target_cod, str) and target_cod.startswith("TEMP_")
                                temp_row_resolved = None
                                if is_temp_id:
                                    try: temp_row_resolved = int(target_cod.split("_")[1])
                                    except ValueError: pass

                                while True:
                                    nombre_val = ws.cell(row=row, column=nombre_col).value
                                    cod_val = ws.cell(row=row, column=id_col).value
                                    if nombre_val and any(x in str(nombre_val).upper() for x in ["TOTAL", "SUMA"]):
                                        break
                                    if nombre_val is None and cod_val is None:
                                        break
                                    if temp_row_resolved == row:
                                        found_row = row
                                        break
                                    elif not is_temp_id and cod_val is not None and str(cod_val).strip() == str(target_cod).strip():
                                        found_row = row
                                        break
                                    row += 1

                                if found_row:
                                    nombre_val = ws.cell(row=found_row, column=nombre_col).value
                                    
                                    # Resolve date for this change
                                    custom_date_str = change_item.get("date") or change_item.get("fecha")
                                    target_date_str = None
                                    if custom_date_str:
                                        custom_date = parse_date_robust(custom_date_str)
                                        if custom_date:
                                            target_date_str = custom_date.strftime("%Y-%m-%d")
                                    
                                    if not target_date_str:
                                        if start_date <= today_mex <= end_date:
                                            target_date_str = today_mex.strftime("%Y-%m-%d")
                                        else:
                                            target_date_str = start_date.strftime("%Y-%m-%d")

                                    existing_inc = {
                                        "faltas": 0,
                                        "retardos": 0,
                                        "vacaciones": 0,
                                        "descuento_adicional": 0.0,
                                        "puntualidad": "SI",
                                        "asistencia": "SI",
                                        "observaciones": "",
                                        "forzar_asistencia": "NO",
                                        "forzar_puntualidad": "NO",
                                        "forzar_vales": "NO",
                                        "ajuste_vales": None,
                                        "ajuste_fondo_ahorro": None
                                    }
                                    for col in schema.get("columns", []):
                                        if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                                            existing_inc[col.get("field")] = 0.0

                                    if ws_inc:
                                        max_c = ws_inc.max_column
                                        for r in range(2, ws_inc.max_row + 1):
                                            r_date = ws_inc.cell(row=r, column=1).value
                                            r_cod = ws_inc.cell(row=r, column=2).value
                                            if r_date and r_cod and str(r_date)[:10] == target_date_str and clean_employee_id(r_cod) == clean_employee_id(target_cod):
                                                existing_inc["faltas"] = int(ws_inc.cell(row=r, column=4).value or 0)
                                                existing_inc["retardos"] = int(ws_inc.cell(row=r, column=5).value or 0)
                                                existing_inc["vacaciones"] = int(ws_inc.cell(row=r, column=6).value or 0)
                                                existing_inc["descuento_adicional"] = float(ws_inc.cell(row=r, column=7).value or 0.0)
                                                existing_inc["puntualidad"] = ws_inc.cell(row=r, column=8).value or "SI"
                                                existing_inc["asistencia"] = ws_inc.cell(row=r, column=9).value or "SI"
                                                existing_inc["observaciones"] = ws_inc.cell(row=r, column=10).value or ""
                                                existing_inc["forzar_asistencia"] = str(ws_inc.cell(row=r, column=11).value or "NO") if max_c >= 11 else "NO"
                                                existing_inc["forzar_puntualidad"] = str(ws_inc.cell(row=r, column=12).value or "NO") if max_c >= 12 else "NO"
                                                existing_inc["forzar_vales"] = str(ws_inc.cell(row=r, column=13).value or "NO") if max_c >= 13 else "NO"
                                                existing_inc["ajuste_vales"] = ws_inc.cell(row=r, column=14).value if max_c >= 14 else None
                                                existing_inc["ajuste_fondo_ahorro"] = ws_inc.cell(row=r, column=15).value if max_c >= 15 else None
                                                
                                                col_idx = 16
                                                for col in schema.get("columns", []):
                                                    if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                                                        existing_inc[col.get("field")] = float(ws_inc.cell(row=r, column=col_idx).value or 0.0)
                                                        col_idx += 1
                                                break
                                    
                                    # Apply AI changes over existing_inc
                                    if "faltas" in change_item:
                                        existing_inc["faltas"] = int(change_item["faltas"])
                                    if "retardos" in change_item:
                                        existing_inc["retardos"] = int(change_item["retardos"])
                                    if "vacaciones" in change_item:
                                        existing_inc["vacaciones"] = int(change_item["vacaciones"])
                                    if "descuento_adicional" in change_item:
                                        existing_inc["descuento_adicional"] = float(change_item["descuento_adicional"])
                                    if "observaciones" in change_item:
                                        existing_inc["observaciones"] = str(change_item["observaciones"])
                                        
                                    if "puntualidad" in change_item:
                                        ai_punt = change_item["puntualidad"]
                                        existing_inc["puntualidad"] = "SI" if (ai_punt is True or str(ai_punt).upper() == "SI") else "NO"
                                    if "asistencia" in change_item:
                                        ai_asist = change_item["asistencia"]
                                        existing_inc["asistencia"] = "SI" if (ai_asist is True or str(ai_asist).upper() == "SI") else "NO"
                                        
                                    if "forzar_asistencia" in change_item:
                                        ai_forzar_asist = change_item["forzar_asistencia"]
                                        existing_inc["forzar_asistencia"] = "SI" if (ai_forzar_asist is True or str(ai_forzar_asist).upper() == "SI") else "NO"
                                    if "forzar_puntualidad" in change_item:
                                        ai_forzar_punt = change_item["forzar_puntualidad"]
                                        existing_inc["forzar_puntualidad"] = "SI" if (ai_forzar_punt is True or str(ai_forzar_punt).upper() == "SI") else "NO"
                                    if "forzar_vales" in change_item:
                                        ai_forzar_val = change_item["forzar_vales"]
                                        existing_inc["forzar_vales"] = "SI" if (ai_forzar_val is True or str(ai_forzar_val).upper() == "SI") else "NO"
                                        
                                    if "ajuste_vales" in change_item:
                                        ai_aj = change_item["ajuste_vales"]
                                        existing_inc["ajuste_vales"] = float(ai_aj) if ai_aj not in ["", None] else None
                                    if "ajuste_fondo_ahorro" in change_item:
                                        ai_fa = change_item["ajuste_fondo_ahorro"]
                                        existing_inc["ajuste_fondo_ahorro"] = float(ai_fa) if ai_fa not in ["", None] else None
                                        
                                    for col in schema.get("columns", []):
                                        if col.get("category") == "deduction" and col.get("incidence_editable") and col.get("field") != "descuento_adicional":
                                            f_name = col.get("field")
                                            if f_name in change_item:
                                                existing_inc[f_name] = float(change_item[f_name])

                                    proposed_changes.append({
                                        "date": target_date_str,
                                        "id": target_cod,
                                        "nombre": nombre_val or "",
                                        **existing_inc
                                    })
                            
                            wb.close()
                            applied_changes = False
                except PermissionError as pe:
                    try: wb.close()
                    except: pass
                    raise pe
                except Exception as parse_e:
                    print("Error parsing and applying changes from AI:", parse_e)
                    try: wb.close()
                    except: pass

                self.send_json({
                    "response": text,
                    "rules_source": rules_source,
                    "offline": False,
                    "applied_changes": applied_changes,
                    "proposed_changes": proposed_changes,
                    "filter_employee_ids": filter_employee_ids
                })
            except PermissionError as perm_e:
                print("Excel file is locked during explain_payroll:", perm_e)
                self.send_json({"error": f"El archivo '{os.path.basename(excel_path)}' estÃ¡ abierto en Microsoft Excel o bloqueado por el sistema. Por favor, cierra el archivo local e intÃ©ntalo de nuevo."}, 500)
                return
            except Exception as e:
                import urllib.error
                print("AI API call failed, falling back to local:", e)
                error_msg = str(e)
                if isinstance(e, urllib.error.HTTPError):
                    try:
                        error_msg = f"HTTP {e.code}: {e.read().decode('utf-8')}"
                    except:
                        pass
                
                customized_desglose = local_desglose + f"\n\n*(Error de API: {error_msg})*"
                self.send_json({"response": customized_desglose, "rules_source": rules_source, "offline": True, "error_details": error_msg})
        except PermissionError:
            self.send_json({"error": f"El archivo '{os.path.basename(excel_path)}' estÃ¡ abierto en Microsoft Excel o bloqueado por el sistema. Por favor, cierra el archivo local e intÃ©ntalo de nuevo."}, 500)
        except Exception as e:
            tb = traceback.format_exc()
            print("Error in explain_payroll endpoint:\n", tb)
            self.send_json({"error": f"Error explaining payroll: {e}", "details": tb}, 500)

if __name__ == "__main__":
    # Ensure users database is initialized/seeded
    load_users()
    socketserver.ThreadingTCPServer.allow_reuse_address = True
    with socketserver.ThreadingTCPServer(("0.0.0.0", PORT), APIHandler) as httpd:
        print(f"Serving RHM CRM & Prenómina on port {PORT}...")
        httpd.serve_forever()
